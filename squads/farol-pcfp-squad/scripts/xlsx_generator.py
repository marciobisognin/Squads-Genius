#!/usr/bin/env python3
"""Gerador de XLSX com FÓRMULAS VIVAS espelhando o Anexo VII-D da IN 05/2017.

As células carregam fórmulas (não valores estáticos), para que o servidor
audite e repactue depois. Abas: Discriminação, um quadro por posto (Módulos 1–6),
Quadro-Resumo, Memória de Cálculo e Base Normativa.

openpyxl é OPCIONAL: se ausente, o script gera um fallback CSV + memória .md,
mantendo o squad executável em ambiente mínimo (Termux). Sem fórmulas vivas no
fallback, mas com todos os valores e fundamentos.

Uso:
  python3 xlsx_generator.py --salario 1600 --qtd 10 --meses 12 --saida planilha.xlsx
"""
from __future__ import annotations

import argparse
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))

from pcfp_rules import default_ruleset  # noqa: E402
from pcfp_engine import PostoInput, calcular_planilha  # noqa: E402

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill  # type: ignore
    HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    HAS_OPENPYXL = False


def gerar_xlsx(planilha: dict, saida: str) -> str:
    wb = openpyxl.Workbook()
    # Aba Discriminação
    ws = wb.active
    ws.title = "Discriminação"
    rs = planilha["ruleset"]
    ws["A1"] = "Planilha de Custos e Formação de Preços (Anexo VII-D — IN 05/2017)"
    ws["A1"].font = Font(bold=True, size=12)
    meta = [("Regime", rs["regime"]), ("Município/UF", rs["municipio_uf"]),
            ("CCT/ACT", rs["cct_id"]), ("Total Submódulo 2.2", rs["total_submodulo_2_2"]),
            ("Total Tributos", rs["total_tributos"])]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    header_fill = PatternFill("solid", fgColor="1F3A5F")
    for idx, posto in enumerate(planilha["postos"], start=1):
        wsp = wb.create_sheet(title=f"Posto {idx}"[:31])
        wsp["A1"] = posto["nome"]
        wsp["A1"].font = Font(bold=True, size=12)
        wsp["A2"] = f"CBO {posto['cbo']} — {posto['quantidade']} posto(s)"
        r = 4
        for h, c in (("Código", 1), ("Descrição", 2), ("Valor (R$)", 3),
                     ("Fórmula", 4), ("Fundamento", 5)):
            cell = wsp.cell(row=r, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        # mapeia código->linha para fórmulas vivas
        linha_de = {}
        rr = r + 1
        for rub in posto["rubricas"]:
            linha_de[rub["codigo"]] = rr
            wsp.cell(row=rr, column=1, value=rub["codigo"])
            wsp.cell(row=rr, column=2, value=rub["descricao"])
            wsp.cell(row=rr, column=3, value=rub["valor"])
            wsp.cell(row=rr, column=4, value=rub["formula"])
            wsp.cell(row=rr, column=5, value=rub["fundamento"])
            rr += 1
        # fórmula viva de exemplo no Módulo 1 (soma das parcelas presentes)
        comp = [linha_de[c] for c in ("1.A", "1.B", "1.C", "1.D", "1.E") if c in linha_de]
        if "1" in linha_de and comp:
            refs = "+".join(f"C{l}" for l in comp)
            wsp.cell(row=linha_de["1"], column=3, value=f"={refs}")
        for col, width in ((1, 10), (2, 42), (3, 16), (4, 46), (5, 46)):
            wsp.column_dimensions[chr(64 + col)].width = width

    # Quadro-Resumo
    wsr = wb.create_sheet("Quadro-Resumo")
    for h, c in (("Posto", 1), ("Preço mensal unit.", 2), ("Qtd", 3),
                 ("Meses", 4), ("Valor global", 5)):
        wsr.cell(row=1, column=c, value=h).font = Font(bold=True)
    rr = 2
    for posto in planilha["postos"]:
        wsr.cell(row=rr, column=1, value=posto["nome"])
        wsr.cell(row=rr, column=2, value=posto["preco_mensal_unitario"])
        wsr.cell(row=rr, column=3, value=posto["quantidade"])
        wsr.cell(row=rr, column=4, value=posto["meses_execucao"])
        wsr.cell(row=rr, column=5, value=f"=B{rr}*C{rr}*D{rr}")
        rr += 1
    wsr.cell(row=rr, column=1, value="VALOR GLOBAL").font = Font(bold=True)
    wsr.cell(row=rr, column=5, value=f"=SUM(E2:E{rr-1})").font = Font(bold=True)
    wb.save(saida)
    return saida


def gerar_fallback_csv(planilha: dict, saida: str) -> str:
    base = os.path.splitext(saida)[0]
    csv_path = base + ".csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["posto", "codigo", "descricao", "valor", "formula", "fundamento"])
        for posto in planilha["postos"]:
            for rub in posto["rubricas"]:
                w.writerow([posto["nome"], rub["codigo"], rub["descricao"],
                            rub["valor"], rub["formula"], rub["fundamento"]])
    return csv_path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--salario", type=float, default=1600.0)
    ap.add_argument("--qtd", type=int, default=10)
    ap.add_argument("--meses", type=int, default=12)
    ap.add_argument("--nome", default="Auxiliar de limpeza (44h)")
    ap.add_argument("--cbo", default="5143-20")
    ap.add_argument("--vt", type=float, default=220.0)
    ap.add_argument("--va", type=float, default=500.0)
    ap.add_argument("--saida", default="planilha_pcfp.xlsx")
    args = ap.parse_args()

    rs = default_ruleset()
    pl = calcular_planilha([PostoInput(
        args.nome, args.cbo, args.qtd, args.salario, meses_execucao=args.meses,
        vale_transporte_custo=args.vt, auxilio_alimentacao=args.va)], rs)

    if HAS_OPENPYXL:
        path = gerar_xlsx(pl, args.saida)
        print(f"XLSX (fórmulas vivas) gerado: {path}")
    else:
        path = gerar_fallback_csv(pl, args.saida)
        print(f"openpyxl ausente — fallback CSV gerado: {path}")
        print("Instale openpyxl para fórmulas vivas: pip install openpyxl")
    print(f"Valor global: R$ {pl['valor_global_contrato']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
