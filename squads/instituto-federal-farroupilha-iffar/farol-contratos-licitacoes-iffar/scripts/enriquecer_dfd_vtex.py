#!/usr/bin/env python3
"""Camada de triangulação de varejo (VTEX) para o Farol Contratos & Licitações IFFar.

Fluxo:
1. Executa a auditoria DFD (`analisar_dfd.py`) — ou recebe uma planilha já
   auditada/enriquecida via `--planilha-auditada` para empilhar as colunas de
   varejo sobre a versão Compras.gov.
2. Pesquisa cada item nas lojas VTEX configuradas (`vtex_catalog.py`).
3. Acrescenta colunas-resumo de varejo à planilha auditada e uma aba de
   evidências com loja, URL e data/hora de cada cotação (IN 65/2021, art. 5º, IV).
4. Gera relatório complementar apenas com as exceções (alertas e revisões).

O varejo é sempre FONTE COMPLEMENTAR: alerta automático só nasce de match de
confiança alta; matches médios/baixos entram como indicativos para revisão
humana. A mediana Compras.gov permanece a referência primária.
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from farol_common import norm, num
from vtex_catalog import DEFAULT_STORES, classify_varejo, parse_lojas, pesquisar_planilha

ROOT = Path(__file__).resolve().parents[1]
ANALYZER = ROOT / "scripts" / "analisar_dfd.py"

EVIDENCE_SHEET = "Evidencias VTEX IN65"

HEADERS = [
    "Varejo VTEX Ofertas Comparáveis",
    "Varejo VTEX Preço Mínimo",
    "Varejo VTEX Preço Mediano",
    "Varejo VTEX Melhor Produto",
    "Varejo VTEX Similaridade Descrição",
    "Varejo VTEX Confiança do Match",
    "Avaliação Preço x Varejo (fonte complementar)",
]

EVIDENCE_HEADERS = [
    "Linha DFD", "Código", "Descrição DFD", "Loja", "Produto", "SKU", "Vendedor",
    "Preço (R$)", "Disponível", "Similaridade", "Link do Produto",
    "URL da Consulta", "Data/Hora da Consulta",
]


def run(cmd: List[str]) -> None:
    print("+", " ".join(str(c) for c in cmd), file=sys.stderr)
    subprocess.run(cmd, check=True)


def find_header(ws) -> int:
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for r in range(1, min(max_row, 40) + 1):
        vals = [str(ws.cell(r, c).value or "").upper() for c in range(1, min(max_col, 80) + 1)]
        if any("DESCRI" in v for v in vals):
            return r
    return 1


def find_col(ws, header: int, labels: List[str]) -> int | None:
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = str(ws.cell(header, c).value or "").upper()
        if any(label in v for label in labels):
            return c
    return None


def index_resumos(resumos: List[Dict[str, Any]]) -> tuple[Dict[int, Dict], Dict[str, Dict]]:
    """Indexa o resumo de varejo por código CATMAT e por descrição normalizada."""
    by_code: Dict[int, Dict] = {}
    by_desc: Dict[str, Dict] = {}
    for r in resumos:
        if r.get("codigo") is not None:
            by_code.setdefault(int(r["codigo"]), r)
        by_desc.setdefault(norm(r.get("descricao")), r)
    return by_code, by_desc


def enrich_workbook(audited_path: Path, resumos: List[Dict[str, Any]],
                    evidencias: List[Dict[str, Any]], out_path: Path) -> Dict[str, Any]:
    by_code, by_desc = index_resumos(resumos)
    wb = openpyxl.load_workbook(audited_path)
    ws = wb[wb.sheetnames[0]]
    header = find_header(ws)
    code_col = find_col(ws, header, ["CÓDIGO", "CODIGO"])
    desc_col = find_col(ws, header, ["DESCRI"])
    price_col = find_col(ws, header, ["VALOR ESTIMADO NA ÚLTIMA", "VALOR ESTIMADO"])
    start = (ws.max_column or 1) + 1
    fill = PatternFill("solid", fgColor="FF1F4E79")
    for i, h in enumerate(HEADERS):
        cell = ws.cell(header, start + i, h)
        cell.fill = fill
        cell.font = Font(color="FFFFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[cell.column_letter].width = 44 if i == 3 else 26
    enriched = 0
    alerts = 0
    review = 0
    for r in range(header + 1, (ws.max_row or header) + 1):
        rec = None
        if code_col:
            try:
                rec = by_code.get(int(float(ws.cell(r, code_col).value)))
            except Exception:
                rec = None
        if rec is None and desc_col:
            rec = by_desc.get(norm(ws.cell(r, desc_col).value))
        if not rec or not rec.get("ofertas"):
            continue
        enriched += 1
        estimado = num(ws.cell(r, price_col).value) if price_col else None
        mediana = num(rec.get("mediana"))
        confianca = rec.get("confianca") or "sem_match"
        avaliacao = classify_varejo(estimado, mediana, confianca)
        if "Acima" in avaliacao or "Abaixo" in avaliacao:
            alerts += 1
        elif "Indicativo" in avaliacao:
            review += 1
        vals = [
            rec.get("comparaveis"),
            rec.get("min"),
            rec.get("mediana"),
            rec.get("melhor_produto"),
            (f"{rec.get('melhor_similaridade', 0):.0%}" if rec.get("melhor_similaridade") is not None else ""),
            confianca,
            avaliacao,
        ]
        for i, val in enumerate(vals):
            cell = ws.cell(r, start + i, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Aba de evidências: uma linha por cotação, com os metadados que a
    # IN 65/2021 exige para uso de e-commerce como fonte de pesquisa de preços.
    if EVIDENCE_SHEET in wb.sheetnames:
        del wb[EVIDENCE_SHEET]
    ev = wb.create_sheet(EVIDENCE_SHEET)
    for c, h in enumerate(EVIDENCE_HEADERS, start=1):
        cell = ev.cell(1, c, h)
        cell.fill = fill
        cell.font = Font(color="FFFFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ev.column_dimensions[cell.column_letter].width = 30
    linha_desc = {r.get("linha"): r.get("descricao") for r in resumos}
    for i, e in enumerate(sorted(evidencias, key=lambda x: (x.get("linha_dfd") or 0, -(x.get("similaridade") or 0))), start=2):
        row = [
            e.get("linha_dfd"), e.get("codigo_dfd"), linha_desc.get(e.get("linha_dfd")),
            e.get("loja"), e.get("produto"), e.get("sku"), e.get("vendedor"),
            e.get("preco"), "sim" if e.get("disponivel") else "não",
            e.get("similaridade"), e.get("link_produto"),
            e.get("url_consulta"), e.get("consultado_em"),
        ]
        for c, val in enumerate(row, start=1):
            ev.cell(i, c, val)
    wb.save(out_path)
    return {
        "itens_enriquecidos": enriched,
        "alertas_varejo": alerts,
        "revisao_humana_match": review,
        "cotacoes_evidencia": len(evidencias),
        "saida": str(out_path),
    }


def write_report(outdir: Path, resumos: List[Dict[str, Any]], enrichment: Dict[str, Any],
                 lojas: List[str], summary_pesquisa: Dict[str, Any]) -> Path:
    path = outdir / "relatorio_varejo_vtex.md"
    excecoes = []
    for r in resumos:
        conf = r.get("confianca")
        if not r.get("ofertas") or conf == "sem_match":
            continue
        excecoes.append(r)
    with path.open("w", encoding="utf-8") as f:
        f.write("# Relatório complementar — Triangulação de varejo (API pública de catálogo VTEX)\n\n")
        f.write(f"Lojas consultadas: **{', '.join(lojas)}**\n\n")
        f.write("Fonte complementar de pesquisa de preços nos termos da IN SEGES/ME nº 65/2021, "
                "art. 5º, inciso IV. Cada cotação registra loja, URL e data/hora da consulta "
                "na aba de evidências da planilha auditada.\n\n")
        f.write("## Cobertura\n\n")
        f.write(f"- Itens pesquisados: {summary_pesquisa.get('itens_pesquisados')}\n")
        f.write(f"- Itens com oferta de varejo encontrada: {summary_pesquisa.get('itens_com_oferta')}\n")
        f.write(f"- Itens enriquecidos na planilha: {enrichment.get('itens_enriquecidos')}\n")
        f.write(f"- Cotações registradas como evidência: {enrichment.get('cotacoes_evidencia')}\n")
        f.write(f"- Alertas de preço (match confiança alta): {enrichment.get('alertas_varejo')}\n")
        f.write(f"- Comparações indicativas aguardando revisão humana do match: {enrichment.get('revisao_humana_match')}\n\n")
        f.write("## Exceções e itens para atenção\n\n")
        listados = 0
        for r in excecoes:
            mediana = r.get("mediana")
            if mediana is None:
                continue
            linha = (f"- Linha {r.get('linha')} — {str(r.get('descricao'))[:90]}: mediana varejo "
                     f"R$ {mediana:,.2f} ({r.get('comparaveis')} cotações, confiança {r.get('confianca')}, "
                     f"melhor match: {str(r.get('melhor_produto'))[:70]} em {r.get('melhor_loja')})\n")
            f.write(linha)
            listados += 1
        if not listados:
            f.write("Nenhum item com comparação de varejo relevante neste ciclo.\n")
        f.write("\n## Limitações e uso responsável\n\n")
        f.write("- Preço de varejo B2C (unitário, com margem e logística embutidas) não é diretamente "
                "comparável a preço de licitação em volume; usar como referência contextual.\n")
        f.write("- A API de catálogo VTEX é o endpoint interno do storefront das lojas: aberta por "
                "design, mas sem contrato formal de estabilidade — pode mudar sem aviso.\n")
        f.write("- Alertas automáticos exigem match de confiança alta; os demais casos requerem "
                "revisão humana da equivalência entre a especificação do DFD e o produto de varejo.\n")
        f.write("- A mediana Compras.gov permanece a referência primária de preço estimado.\n\n")
        f.write("Licença: MIT. Criado por Marcio Bisognin. Instagram: @marciobisognin.\n")
    return path


def main() -> int:
    ap = argparse.ArgumentParser(description="Audita DFD e triangula preços com lojas de varejo VTEX (fonte complementar IN 65/2021).")
    ap.add_argument("planilha")
    ap.add_argument("--out", default="output/farol-varejo-vtex")
    ap.add_argument("--lojas", help="domínios VTEX separados por vírgula (padrão: %s)" % ", ".join(DEFAULT_STORES))
    ap.add_argument("--quantidade", type=int, default=20, help="ofertas por loja e item (máx. 50)")
    ap.add_argument("--sleep", type=float, default=0.2)
    ap.add_argument("--max-itens", type=int, help="limita itens pesquisados para teste/demonstração")
    ap.add_argument("--planilha-auditada", help="planilha já auditada/enriquecida (ex.: *_AUDITADA_COMPRAS_GOV.xlsx) para empilhar o varejo sem repetir a auditoria")
    args = ap.parse_args()

    outdir = Path(args.out)
    outdir.mkdir(parents=True, exist_ok=True)
    lojas = parse_lojas(args.lojas)

    if args.planilha_auditada:
        audited_path = Path(args.planilha_auditada)
        audit_summary: Dict[str, Any] = {"planilha_auditada_reaproveitada": str(audited_path)}
    else:
        audit_dir = outdir / "01_auditoria_dfd"
        run([sys.executable, str(ANALYZER), args.planilha, "--out", str(audit_dir)])
        audit_summary = json.loads((audit_dir / "summary.json").read_text(encoding="utf-8"))
        audited_path = Path(audit_summary["outputs"]["planilha"])

    varejo_dir = outdir / "02_varejo_vtex"
    cache_dir = outdir / ".cache"
    summary_pesquisa = pesquisar_planilha(args.planilha, lojas, varejo_dir,
                                          quantidade=args.quantidade, sleep=args.sleep,
                                          max_itens=args.max_itens, cache_dir=cache_dir)
    resumos = json.loads((varejo_dir / "resumo_varejo_por_item.json").read_text(encoding="utf-8"))
    evidencias = json.loads((varejo_dir / "evidencias_varejo.json").read_text(encoding="utf-8"))

    enriched_path = outdir / (Path(args.planilha).stem + "_AUDITADA_VAREJO_VTEX.xlsx")
    enrichment = enrich_workbook(audited_path, resumos, evidencias, enriched_path)
    report_path = write_report(outdir, resumos, enrichment, lojas, summary_pesquisa)

    summary = {
        "planilha_original": args.planilha,
        "lojas_vtex": lojas,
        "auditoria_dfd": audit_summary,
        "pesquisa_varejo": summary_pesquisa,
        "enriquecimento": enrichment,
        "relatorio_varejo_vtex": str(report_path),
    }
    (outdir / "summary_varejo_vtex.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
