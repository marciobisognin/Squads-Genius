#!/usr/bin/env python3
"""Auditoria de planilha DFD/lista de itens do squad Farol Contratos & Licitações IFFar."""
import argparse, csv, html, json, math, re, statistics
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from farol_common import norm, num, date_range_default, compras_fetch_material_precos

RISK_COLOR = {'ALTO':'FFFF9999','MÉDIO':'FFFFE699','BAIXO':'FFD9EAD3','OK':'FFE2F0D9'}
PACK_TERMS = ['CAIXA','PACOTE','FARDO','KIT','CONJUNTO','JOGO','PAR','ROLO','FRASCO','POTE','GALÃO','SACO','EMBALAGEM','CARTELA']
RESTRICTIVE = ['MARCA','MODELO EXCLUSIVO','FABRICAÇÃO NACIONAL','NACIONAL']
# Frases que indicam marca citada apenas como referência de padrão de qualidade,
# com similar/equivalente admitido — uso permitido pelo art. 41 da Lei nº
# 14.133/2021. Sem essa exceção, qualquer menção a "marca" era sempre apontada
# como termo restritivo, mesmo quando a redação já está em conformidade.
BRAND_REFERENCE_SAFE = re.compile(
    r'MARCA\s+DE\s+REFER[ÊE]NCIA'
    r'|A\s+T[ÍI]TULO\s+DE\s+REFER[ÊE]NCIA'
    r'|ADMITID[OA]S?\s+(?:PRODUTOS?\s+)?SIMILAR'
    r'|OU\s+SIMILAR'
    r'|OU\s+EQUIVALENTE'
    r'|QUALIDADE\s+EQUIVALENTE'
    r'|N[ÃA]O\s+SER[ÁA]\s+ACEITA\s+INDICA[ÇC][ÃA]O\s+DE\s+MARCA'
)
TYPO = {'CADO':'CABO','INXIDÁVEL':'INOXIDÁVEL','M`NIMA':'MÍNIMA','MINIMA':'MÍNIMA','CARACTERISTICAS':'CARACTERÍSTICAS','ACO ':'AÇO '}

# Perfil padrão de mapeamento de colunas. Pode ser sobrescrito com --perfil perfil.json
DEFAULT_PROFILE = {
    'codigo': ['CÓDIGO', 'CODIGO'],
    'descricao': ['DESCRIÇÃO', 'DESCRICAO'],
    'unidade': ['UNIDADE'],
    'preco': ['VALOR ESTIMADO NA ÚLTIMA', 'VALOR ESTIMADO'],
    'quantidade': ['QUANTIDADE ESTIMADA'],
    'valor_total': ['VALOR TOTAL'],
    'header': ['DESCRIÇÃO', 'UNIDADE'],
    'campus_row': 2,
}


def load_profile(path=None):
    profile = dict(DEFAULT_PROFILE)
    if path:
        profile.update(json.loads(Path(path).read_text(encoding='utf-8')))
    return profile

def find_header(ws, profile=None):
    profile = profile or DEFAULT_PROFILE
    kw_desc, kw_unid = profile['header'][0], profile['header'][1]
    for r in range(1, min(ws.max_row,30)+1):
        vals=[norm(ws.cell(r,c).value) for c in range(1, ws.max_column+1)]
        if any(kw_desc in v for v in vals) and any(kw_unid in v for v in vals):
            return r
    return 1

def campus_for_col(ws, col, header_row, campus_row=2):
    # prefer the nearest non-empty cell on the campus row to the left within a 2-column campus block
    for c in [col, col-1, col-2]:
        if c >= 1:
            v = ws.cell(campus_row,c).value
            if v and 'DOCUMENTO DE FORMALIZAÇÃO' not in str(v).upper() and 'VALOR' not in str(v).upper():
                return str(v).strip()
    return ws.cell(header_row,col).column_letter

def is_yellow(cell):
    rgb = getattr(cell.fill.fgColor, 'rgb', None)
    if not isinstance(rgb, str):
        return False
    rgb = rgb.upper()
    if rgb in ('FFFFFF00','FFFF00','00FFFF00'):
        return True
    if len(rgb) == 8:
        try:
            r,g,b = int(rgb[2:4],16), int(rgb[4:6],16), int(rgb[6:8],16)
        except ValueError:
            return False
        return r >= 200 and g >= 190 and b <= 130
    return False

def detect_columns(ws, header_row, profile=None):
    profile = profile or DEFAULT_PROFILE
    cols={}
    valor_total_cols=[]
    for c in range(1, ws.max_column+1):
        h=norm(ws.cell(header_row,c).value)
        if not h: continue
        if 'codigo' not in cols and any(k in h for k in profile['codigo']): cols['codigo']=c
        elif 'descricao' not in cols and any(k in h for k in profile['descricao']): cols['descricao']=c
        elif 'unidade' not in cols and any(k in h for k in profile['unidade']): cols['unidade']=c
        elif 'preco' not in cols and any(k in h for k in profile['preco']) and 'TOTAL' not in h: cols['preco']=c
        if any(k in h for k in profile['valor_total']): valor_total_cols.append(c)
    # só usa coluna de valor total quando ela é inequívoca (uma única na planilha)
    if len(valor_total_cols) == 1:
        cols['valor_total']=valor_total_cols[0]
    qcols=[]
    campus_row=profile.get('campus_row',2)
    for c in range(1, ws.max_column+1):
        h=norm(ws.cell(header_row,c).value)
        if any(k in h for k in profile['quantidade']):
            sample=ws.cell(header_row+1,c)
            if is_yellow(sample) or is_yellow(ws.cell(header_row,c)):
                qcols.append((c, campus_for_col(ws,c,header_row,campus_row)))
    if not qcols:
        # fallback: aceita todas as colunas de quantidade quando não há marcação amarela
        for c in range(1, ws.max_column+1):
            h=norm(ws.cell(header_row,c).value)
            if any(k in h for k in profile['quantidade']):
                qcols.append((c, campus_for_col(ws,c,header_row,campus_row)))
    return cols,qcols

def description_findings(desc, unidade):
    d=norm(desc); u=norm(unidade)
    findings=[]
    if len(d) < 45:
        findings.append(('ALTO','DESCRIÇÃO','Descrição muito curta; complementar especificação técnica mínima.'))
    if not re.search(r'\d', d):
        findings.append(('MÉDIO','DESCRIÇÃO','Descrição sem medida/capacidade/dimensão numérica; verificar se o item exige especificação objetiva.'))
    for wrong,right in TYPO.items():
        if wrong in d:
            findings.append(('BAIXO','DESCRIÇÃO',f'Possível erro de digitação: "{wrong}"; sugerir "{right}".'))
    for term in RESTRICTIVE:
        if term in d:
            if term == 'MARCA' and BRAND_REFERENCE_SAFE.search(d):
                continue
            sev='MÉDIO' if term != 'FABRICAÇÃO NACIONAL' else 'ALTO'
            findings.append((sev,'DESCRIÇÃO',f'Termo potencialmente restritivo/direcionador: "{term}"; revisar justificativa ou remover.'))
    if any(t in d for t in ['INOX','ALUMÍNIO','ALUMINIO','VIDRO','PLÁSTICO','PLASTICO','POLIPROPILENO','AÇO','ACO']):
        pass
    else:
        if any(k in d for k in ['PANELA','CAÇAROLA','FORMA','COPO','XÍCARA','XICARA','TALHER','FACA','COLHER','PRATO','GARRAFA','JARRA']):
            findings.append(('MÉDIO','DESCRIÇÃO','Material constitutivo não aparece de forma clara; confirmar especificação.'))
    # unit/package mismatch
    if u in ['UNIDADE','UND','UN'] and any(t in d for t in PACK_TERMS):
        findings.append(('MÉDIO','UNIDADE','Descrição menciona embalagem/conjunto, mas unidade de fornecimento está como unidade; esclarecer se a cotação é por unidade ou embalagem.'))
    if u in ['CAIXA','PACOTE','FARDO'] and not any(t in d for t in PACK_TERMS):
        findings.append(('MÉDIO','UNIDADE',f'Unidade "{u}" exige quantidade por embalagem clara na descrição.'))
    return findings

def outlier_findings(values_by_campus):
    vals=[v for c,v in values_by_campus if v is not None and v>0]
    out=[]
    if len(vals) < 4:
        return out
    med=statistics.median(vals)
    qs=statistics.quantiles(vals, n=4, method='inclusive') if len(vals)>=4 else [med,med,med]
    q1,q3=qs[0],qs[2]
    iqr=max(q3-q1,0)
    mad=statistics.median([abs(v-med) for v in vals]) if vals else 0
    for campus,v in values_by_campus:
        if v is None: continue
        if v > 0:
            high = (iqr>0 and v>q3+1.5*iqr and v>=2.5*max(med,1)) or (mad>0 and abs(v-med)/(1.4826*mad)>3.5 and v>med)
            low = (med>=4 and v<=0.25*med and v<q1-1.5*iqr) or (mad>0 and abs(v-med)/(1.4826*mad)>3.5 and v<med)
            if high:
                out.append(('ALTO','OUTLIER',f'{campus} solicitou {v:g}, acima da mediana {med:g}; confirmar demanda ou possível erro.'))
            elif low:
                out.append(('MÉDIO','OUTLIER',f'{campus} solicitou {v:g}, abaixo da mediana {med:g}; confirmar interpretação do item.'))
    zeros=[campus for campus,v in values_by_campus if v == 0]
    positives=sum(1 for c,v in values_by_campus if v and v>0)
    if positives >= max(6, math.ceil(0.7*len(values_by_campus))) and zeros and len(zeros)<=3:
        out.append(('BAIXO','OUTLIER',f'Campus sem demanda em item com adesão ampla: {", ".join(zeros[:5])}; confirmar se zero é intencional.'))
    return out

def price_findings(price):
    if price is None: return [('MÉDIO','PREÇO','Preço estimado ausente; realizar pesquisa/validação.')]
    if price <= 0: return [('ALTO','PREÇO','Preço estimado zero ou negativo; corrigir antes da licitação.')]
    return []

def total_consistency_findings(price, qty_total, declared_total):
    """Compara valor total declarado na planilha com preço unitário × quantidade total."""
    if not price or price <= 0 or not qty_total or declared_total is None or declared_total <= 0:
        return []
    calc = price * qty_total
    if calc <= 0:
        return []
    if abs(calc - declared_total) / max(declared_total, calc) > 0.05:
        return [('MÉDIO','PREÇO',f'Valor total informado R$ {declared_total:,.2f} diverge do calculado (preço × quantidade total = R$ {calc:,.2f}); verificar fórmula ou preenchimento.')]
    return []


def compras_price_findings(price, stats):
    findings=[]
    if not stats:
        return findings
    if stats.get('erro'):
        findings.append(('BAIXO','COMPRAS.GOV',f'Pesquisa Compras.gov não concluída: {stats.get("erro")}.'))
        return findings
    registros=stats.get('registros',0)
    med=stats.get('mediana')
    if registros == 0:
        findings.append(('BAIXO','COMPRAS.GOV','Sem preço praticado encontrado no período pesquisado; avaliar ampliar período/filtros.'))
    if price and med:
        if price > med*1.8:
            findings.append(('ALTO','PREÇO EXTERNO',f'Preço estimado {price:.2f} está acima da mediana Compras.gov {med:.2f}; revisar pesquisa de preços.'))
        elif price < med*0.45:
            findings.append(('MÉDIO','PREÇO EXTERNO',f'Preço estimado {price:.2f} está abaixo da mediana Compras.gov {med:.2f}; confirmar unidade/embalagem e compatibilidade.'))
    return findings


def risk_rank(risks):
    if 'ALTO' in risks: return 'ALTO'
    if 'MÉDIO' in risks: return 'MÉDIO'
    if 'BAIXO' in risks: return 'BAIXO'
    return 'OK'

def analyze(input_path, outdir, pesquisa_compras_gov=False, compras_inicio=None, compras_fim=None, compras_paginas=1, compras_tamanho_pagina=10, perfil=None, cache_dir=None):
    default_inicio, default_fim = date_range_default()
    compras_inicio = compras_inicio or default_inicio
    compras_fim = compras_fim or default_fim
    profile = load_profile(perfil)
    input_path=Path(input_path); outdir=Path(outdir); outdir.mkdir(parents=True, exist_ok=True)
    if pesquisa_compras_gov and cache_dir is None:
        cache_dir = outdir / '.cache'
    wb=openpyxl.load_workbook(input_path)
    ws=wb[wb.sheetnames[0]]
    header=find_header(ws, profile)
    cols,qcols=detect_columns(ws, header, profile)
    required=['codigo','descricao','unidade']
    missing=[k for k in required if k not in cols]
    if missing: raise SystemExit('Colunas obrigatórias ausentes: '+', '.join(missing))
    action_col=ws.max_column+1
    headers=['Ações Necessárias','Nível de Risco','Tipos de Achado','Outliers Quantitativos','Sugestão de Decisão','Valor Total Estimado (R$)']
    if pesquisa_compras_gov:
        headers.extend(['Compras.gov Registros','Compras.gov Mediana','Compras.gov Média','Compras.gov Min/Max'])
    for i,h in enumerate(headers):
        cell=ws.cell(header, action_col+i, h)
        cell.fill=PatternFill('solid', fgColor='FF1F4E78')
        cell.font=Font(color='FFFFFFFF', bold=True)
        cell.alignment=Alignment(wrap_text=True, vertical='top')
    findings=[]; rows=0; counts={'ALTO':0,'MÉDIO':0,'BAIXO':0,'OK':0}
    type_counts={}
    valor_por_risco={'ALTO':0.0,'MÉDIO':0.0,'BAIXO':0.0,'OK':0.0}
    itens_financeiro=[]
    for r in range(header+1, ws.max_row+1):
        desc=ws.cell(r, cols['descricao']).value
        if not desc or not str(desc).strip(): continue
        rows += 1
        codigo=ws.cell(r, cols.get('codigo',1)).value
        item=ws.cell(r, 1).value
        unidade=ws.cell(r, cols.get('unidade')).value
        price=num(ws.cell(r, cols.get('preco')).value) if cols.get('preco') else None
        declared_total=num(ws.cell(r, cols.get('valor_total')).value) if cols.get('valor_total') else None
        vals=[(campus, num(ws.cell(r,c).value)) for c,campus in qcols]
        qty_total=sum(v for _,v in vals if v and v>0)
        valor_estimado=round(price*qty_total,2) if price and price>0 and qty_total else None
        fs=[]
        compras_stats=None
        if pesquisa_compras_gov:
            compras_stats=compras_fetch_material_precos(codigo, compras_inicio, compras_fim, compras_paginas, compras_tamanho_pagina, cache_dir=cache_dir)
        fs.extend(description_findings(desc, unidade))
        fs.extend(price_findings(price))
        fs.extend(total_consistency_findings(price, qty_total, declared_total))
        fs.extend(compras_price_findings(price, compras_stats))
        fs.extend(outlier_findings(vals))
        if not fs:
            actions='Aprovado sem ressalvas'
            risk='OK'; types='OK'; outs=''; decision='Aprovar sem ressalvas automatizadas.'
        else:
            risk=risk_rank([f[0] for f in fs])
            actions=' | '.join(f[2] for f in fs)
            types=', '.join(sorted(set(f[1] for f in fs)))
            outs=' | '.join(f[2] for f in fs if f[1]=='OUTLIER')
            decision='Revisar antes de consolidar.' if risk in ['ALTO','MÉDIO'] else 'Baixa criticidade; revisar se houver tempo.'
        counts[risk]+=1
        if valor_estimado:
            valor_por_risco[risk]+=valor_estimado
            itens_financeiro.append({'linha':r,'item':item,'codigo':codigo,'risco':risk,'valor':valor_estimado,'descricao':str(desc)[:120]})
        for f in fs: type_counts[f[1]]=type_counts.get(f[1],0)+1
        fill=PatternFill('solid', fgColor=RISK_COLOR[risk])
        row_values=[actions,risk,types,outs,decision, valor_estimado if valor_estimado is not None else '']
        if pesquisa_compras_gov:
            if compras_stats and not compras_stats.get('erro'):
                row_values.extend([
                    compras_stats.get('registros',0),
                    round(compras_stats.get('mediana',0),4) if compras_stats.get('mediana') else '',
                    round(compras_stats.get('media',0),4) if compras_stats.get('media') else '',
                    (f"{compras_stats.get('min'):.2f}/{compras_stats.get('max'):.2f}" if compras_stats.get('min') is not None else '')
                ])
            else:
                row_values.extend(['ERRO','','',''])
        for i,val in enumerate(row_values):
            cell=ws.cell(r, action_col+i, val)
            cell.alignment=Alignment(wrap_text=True, vertical='top')
            if i==1: cell.fill=fill
        if fs:
            for sev,typ,msg in fs:
                findings.append({'linha':r,'item':item,'codigo':codigo,'unidade':unidade,'risco':sev,'tipo':typ,'achado':msg,'valor_estimado':valor_estimado or '','descricao':str(desc)[:300]})
    for c in range(action_col, action_col+len(headers)):
        ws.column_dimensions[ws.cell(header,c).column_letter].width = 32 if c==action_col else 22
    audited=outdir/(input_path.stem+'_AUDITADA.xlsx')
    wb.save(audited)
    csv_path=outdir/'achados_auditoria.csv'
    with csv_path.open('w', newline='', encoding='utf-8-sig') as f:
        w=csv.DictWriter(f, fieldnames=['linha','item','codigo','unidade','risco','tipo','achado','valor_estimado','descricao'])
        w.writeheader(); w.writerows(findings)
    report=outdir/'relatorio_executivo.md'
    report.write_text(make_report(input_path, rows, counts, type_counts, findings, qcols, valor_por_risco, itens_financeiro, pesquisa_compras_gov, compras_inicio, compras_fim), encoding='utf-8')
    dash=outdir/'dashboard_auditoria.html'
    dash.write_text(make_dashboard(input_path, rows, counts, type_counts, findings, valor_por_risco), encoding='utf-8')
    summary={'input':str(input_path),'items_analisados':rows,'campi_detectados':[c for _,c in qcols],'achados':len(findings),'riscos':counts,'tipos':type_counts,'valor_estimado_por_risco':{k:round(v,2) for k,v in valor_por_risco.items()},'integracao_compras_gov':{'ativa': bool(pesquisa_compras_gov), 'inicio': compras_inicio, 'fim': compras_fim, 'paginas': compras_paginas},'outputs':{'planilha':str(audited),'csv':str(csv_path),'relatorio':str(report),'dashboard':str(dash)}}
    (outdir/'summary.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    return summary

def brl(v):
    return f'R$ {v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

def make_report(input_path, rows, counts, type_counts, findings, qcols, valor_por_risco, itens_financeiro, pesquisa_compras_gov=False, compras_inicio='', compras_fim=''):
    crit=[f for f in findings if f['risco']=='ALTO'][:15]
    lines=[]
    lines.append(f'# Relatório Executivo — Auditoria de DFD/Listas de Itens\n')
    lines.append(f'**Arquivo analisado:** `{Path(input_path).name}`')
    lines.append(f'**Data de geração:** {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    lines.append(f'**Itens analisados:** {rows}')
    lines.append(f'**Campi/colunas oficiais detectados:** {len(qcols)} — '+', '.join(c for _,c in qcols))
    lines.append('\n## Síntese de risco')
    for k in ['ALTO','MÉDIO','BAIXO','OK']:
        lines.append(f'- {k}: {counts.get(k,0)} itens')
    lines.append('\n## Valor financeiro sob risco')
    lines.append('Estimativa por item: preço unitário × quantidade total multicampi.')
    for k in ['ALTO','MÉDIO','BAIXO','OK']:
        lines.append(f'- {k}: {brl(valor_por_risco.get(k,0.0))}')
    top_fin=sorted([i for i in itens_financeiro if i['risco'] in ('ALTO','MÉDIO')], key=lambda x:-x['valor'])[:10]
    if top_fin:
        lines.append('\n### Maiores valores em itens com risco ALTO/MÉDIO')
        for i in top_fin:
            lines.append(f'- Linha {i["linha"]}, código {i["codigo"]} ({i["risco"]}): {brl(i["valor"])} — {i["descricao"]}')
    lines.append('\n## Achados por tipo')
    for k,v in sorted(type_counts.items(), key=lambda x:-x[1]): lines.append(f'- {k}: {v}')
    lines.append('\n## Itens prioritários')
    if crit:
        for f in crit:
            lines.append(f'- Linha {f["linha"]}, item {f["item"]}, código {f["codigo"]}: [{f["tipo"]}] {f["achado"]}')
    else:
        lines.append('- Não foram detectados achados de risco alto pelas regras automatizadas.')
    lines.append('\n## Integração Compras.gov')
    if pesquisa_compras_gov:
        lines.append(f'- Pesquisa externa ativada no período {compras_inicio} a {compras_fim}.')
        lines.append('- A planilha auditada recebeu colunas com registros, mediana, média e faixa min/max do Compras.gov por código CATMAT.')
        lines.append('- Achados do tipo PREÇO EXTERNO comparam o preço estimado interno com a mediana de preços praticados.')
    else:
        lines.append('- Pesquisa externa não ativada nesta execução. Use `--pesquisa-compras-gov` para enriquecer com preços praticados.')
    lines.append('\n## Recomendações imediatas')
    lines.append('- Revisar primeiro itens com risco ALTO e achados de OUTLIER/PREÇO/UNIDADE, priorizando os de maior valor financeiro.')
    lines.append('- Validar com o campus demandante qualquer quantitativo muito acima ou abaixo da mediana multicampi.')
    lines.append('- Registrar descrições saneadas na base de conhecimento (`scripts/base_conhecimento.py`) para reaproveitamento.')
    lines.append('- Acompanhar pendências no painel de saneamento (`scripts/painel_saneamento.py`).')
    lines.append('- Para itens financeiramente relevantes, acionar pesquisa externa de preços em PNCP/Compras.gov/Painel de Preços.')
    lines.append('\n## Limitações')
    lines.append('Esta análise é apoio técnico automatizado. A equipe de licitações/contratos deve validar juridicamente, tecnicamente e com os campi os achados antes de alterar o DFD ou edital.')
    lines.append('\nLicença: MIT. Criado por Marcio Bisognin. Instagram: @marciobisognin.')
    return '\n'.join(lines)+'\n'

def risk_chart_svg(counts):
    order=['ALTO','MÉDIO','BAIXO','OK']
    colors={'ALTO':'#ef4444','MÉDIO':'#f59e0b','BAIXO':'#a3e635','OK':'#34d399'}
    mx=max([counts.get(k,0) for k in order]+[1])
    bars=[]
    for i,k in enumerate(order):
        v=counts.get(k,0); h=int(150*v/mx)
        x=30+i*115
        bars.append(f'<rect x="{x}" y="{180-h}" width="80" height="{h}" rx="6" fill="{colors[k]}"/>'
                    f'<text x="{x+40}" y="205" text-anchor="middle" fill="#cbd5e1" font-size="13">{k}</text>'
                    f'<text x="{x+40}" y="{172-h}" text-anchor="middle" fill="#e5e7eb" font-size="15">{v}</text>')
    return f'<svg viewBox="0 0 500 220" width="500" height="220" role="img" aria-label="Distribuição de risco">{"".join(bars)}</svg>'

def make_dashboard(input_path, rows, counts, type_counts, findings, valor_por_risco=None):
    valor_por_risco = valor_por_risco or {}
    cards=''.join(f'<div class="card"><b>{html.escape(k)}</b><span>{v}</span><small>{html.escape(brl(valor_por_risco.get(k,0.0)))}</small></div>' for k,v in counts.items())
    type_rows=''.join(f'<tr><td>{html.escape(k)}</td><td>{v}</td></tr>' for k,v in sorted(type_counts.items(), key=lambda x:-x[1]))
    finding_rows=''.join(f'<tr><td>{f["linha"]}</td><td>{html.escape(str(f["codigo"]))}</td><td>{html.escape(f["risco"])}</td><td>{html.escape(f["tipo"])}</td><td>{html.escape(f["achado"])}</td></tr>' for f in findings[:80])
    chart=risk_chart_svg(counts)
    return f"""<!doctype html><html lang=\"pt-BR\"><head><meta charset=\"utf-8\"><title>Dashboard Auditoria DFD</title><style>body{{font-family:Arial,sans-serif;background:#0f172a;color:#e5e7eb;margin:0;padding:28px}}h1{{color:#93c5fd}}.grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}}.card{{background:#1e293b;border:1px solid #334155;border-radius:14px;padding:18px}}.card span{{display:block;font-size:34px;color:#fbbf24}}.card small{{color:#94a3b8}}table{{width:100%;border-collapse:collapse;background:#111827;margin-top:18px}}td,th{{border:1px solid #374151;padding:8px;vertical-align:top}}th{{background:#1f2937}}.muted{{color:#94a3b8}}.chart{{background:#1e293b;border:1px solid #334155;border-radius:14px;padding:18px;margin-top:18px}}</style></head><body><h1>Dashboard — Auditoria de DFD/Listas de Itens</h1><p class=\"muted\">Arquivo: {html.escape(Path(input_path).name)} | Itens analisados: {rows}</p><div class=\"grid\">{cards}</div><div class=\"chart\"><h2>Distribuição de risco</h2>{chart}</div><h2>Achados por tipo</h2><table><tr><th>Tipo</th><th>Quantidade</th></tr>{type_rows}</table><h2>Achados prioritários</h2><table><tr><th>Linha</th><th>Código</th><th>Risco</th><th>Tipo</th><th>Achado</th></tr>{finding_rows}</table><p class=\"muted\">Licença: MIT. Criado por Marcio Bisognin. Instagram: @marciobisognin.</p></body></html>"""

def main():
    inicio_padrao, fim_padrao = date_range_default()
    ap=argparse.ArgumentParser(description='Audita planilha DFD/lista de itens de licitações e contratos.')
    ap.add_argument('planilha')
    ap.add_argument('--out', default='output/auditoria')
    ap.add_argument('--perfil', help='JSON com mapeamento de colunas para formatos de planilha diferentes do padrão IFFar')
    ap.add_argument('--pesquisa-compras-gov', action='store_true', help='enriquece a análise com preços praticados da API Dados Abertos Compras.gov.br')
    ap.add_argument('--compras-inicio', default=inicio_padrao, help='data inicial YYYY-MM-DD para pesquisa Compras.gov (padrão: 24 meses atrás)')
    ap.add_argument('--compras-fim', default=fim_padrao, help='data final YYYY-MM-DD para pesquisa Compras.gov (padrão: hoje)')
    ap.add_argument('--compras-paginas', type=int, default=1, help='páginas por item a consultar no Compras.gov')
    ap.add_argument('--compras-tamanho-pagina', type=int, default=10, help='tamanho da página da API, mínimo 10')
    ap.add_argument('--cache', help='diretório de cache das consultas Compras.gov (padrão: <out>/.cache)')
    args=ap.parse_args()
    summary=analyze(args.planilha, args.out, args.pesquisa_compras_gov, args.compras_inicio, args.compras_fim, args.compras_paginas, args.compras_tamanho_pagina, perfil=args.perfil, cache_dir=args.cache)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
if __name__ == '__main__': main()
