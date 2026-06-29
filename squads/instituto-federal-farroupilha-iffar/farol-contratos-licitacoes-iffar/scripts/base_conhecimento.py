#!/usr/bin/env python3
"""Base de conhecimento de descrições saneadas do squad Farol IFFar.

Guarda descrições revisadas/aprovadas pela equipe de licitações para
reaproveitamento em ciclos futuros e permite verificar uma planilha DFD
contra as descrições já aprovadas.

Uso:
    python scripts/base_conhecimento.py adicionar --codigo 437939 --descricao "..." --unidade UNIDADE
    python scripts/base_conhecimento.py buscar --codigo 437939
    python scripts/base_conhecimento.py buscar --texto "caneta esferográfica"
    python scripts/base_conhecimento.py verificar planilha.xlsx --out output/verificacao_kb.csv
"""
from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from farol_common import similarity

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BASE = ROOT / "knowledge" / "base_descricoes.json"


def load_base(path: Path) -> List[Dict[str, Any]]:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return []


def save_base(path: Path, entries: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")


def cmd_adicionar(args: argparse.Namespace) -> int:
    path = Path(args.base)
    entries = load_base(path)
    entry = {
        "codigo": str(args.codigo),
        "descricao_aprovada": args.descricao,
        "unidade": args.unidade,
        "observacao": args.obs or "",
        "registrado_em": datetime.now().isoformat(timespec="seconds"),
    }
    entries = [e for e in entries if e.get("codigo") != str(args.codigo)] + [entry]
    save_base(path, entries)
    print(json.dumps({"status": "ok", "codigo": str(args.codigo), "total_na_base": len(entries), "base": str(path)}, ensure_ascii=False, indent=2))
    return 0


def cmd_buscar(args: argparse.Namespace) -> int:
    entries = load_base(Path(args.base))
    if args.codigo:
        matches = [e for e in entries if e.get("codigo") == str(args.codigo)]
    elif args.texto:
        scored = [(similarity(args.texto, e.get("descricao_aprovada")), e) for e in entries]
        matches = [dict(e, similaridade=round(s, 3)) for s, e in sorted(scored, key=lambda x: -x[0]) if s > 0][:10]
    else:
        matches = entries
    print(json.dumps({"total_na_base": len(entries), "resultados": matches}, ensure_ascii=False, indent=2))
    return 0


def cmd_verificar(args: argparse.Namespace) -> int:
    from analisar_dfd import find_header, detect_columns

    import openpyxl

    entries = {e["codigo"]: e for e in load_base(Path(args.base))}
    wb = openpyxl.load_workbook(args.planilha, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = find_header(ws)
    cols, _ = detect_columns(ws, header)
    rows: List[Dict[str, Any]] = []
    for r in range(header + 1, ws.max_row + 1):
        desc = ws.cell(r, cols.get("descricao", 3)).value
        if not desc or not str(desc).strip():
            continue
        codigo = ws.cell(r, cols.get("codigo", 2)).value
        try:
            codigo = str(int(float(codigo)))
        except Exception:
            codigo = str(codigo or "")
        kb = entries.get(codigo)
        if not kb:
            continue
        sim = similarity(desc, kb["descricao_aprovada"])
        rows.append({
            "linha": r,
            "codigo": codigo,
            "descricao_planilha": str(desc)[:200],
            "descricao_aprovada": kb["descricao_aprovada"][:200],
            "similaridade": round(sim, 3),
            "situacao": "COMPATÍVEL" if sim >= 0.6 else "DIVERGE DA DESCRIÇÃO APROVADA",
        })
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["linha", "codigo", "descricao_planilha", "descricao_aprovada", "similaridade", "situacao"])
        w.writeheader()
        w.writerows(rows)
    divergentes = sum(1 for r in rows if r["situacao"] != "COMPATÍVEL")
    print(json.dumps({"status": "ok", "itens_na_base": len(rows), "divergentes": divergentes, "csv": str(out_path)}, ensure_ascii=False, indent=2))
    return 0


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="base-conhecimento", description="Base de descrições saneadas para reaproveitamento entre ciclos.")
    p.add_argument("--base", default=str(DEFAULT_BASE), help="caminho do arquivo JSON da base")
    sub = p.add_subparsers(dest="cmd", required=True)
    a = sub.add_parser("adicionar", help="registra/atualiza uma descrição aprovada")
    a.add_argument("--codigo", required=True)
    a.add_argument("--descricao", required=True)
    a.add_argument("--unidade", required=True)
    a.add_argument("--obs")
    a.set_defaults(func=cmd_adicionar)
    b = sub.add_parser("buscar", help="busca por código ou texto livre")
    b.add_argument("--codigo")
    b.add_argument("--texto")
    b.set_defaults(func=cmd_buscar)
    v = sub.add_parser("verificar", help="confere uma planilha DFD contra as descrições aprovadas")
    v.add_argument("planilha")
    v.add_argument("--out", default="output/verificacao_base_conhecimento.csv")
    v.set_defaults(func=cmd_verificar)
    return p


def main() -> int:
    args = build_parser().parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
