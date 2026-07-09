#!/usr/bin/env python3
"""vtex-catalog: coletor de preços de varejo via API pública de catálogo VTEX.

Toda loja construída sobre a plataforma VTEX expõe, sem autenticação, o
endpoint `/api/catalog_system/pub/products/search` usado pelo próprio
frontend. Este CLI consulta esse endpoint em lojas configuráveis e produz
cotações de varejo como FONTE COMPLEMENTAR de pesquisa de preços
(IN SEGES/ME nº 65/2021, art. 5º, inciso IV), registrando para cada oferta
a URL da consulta, a data/hora e a loja — metadados exigidos pela IN.

Limites conhecidos da API (ver references/vtex-catalogo-integracao.md):
- janela máxima de 50 itens por requisição (`_to - _from <= 49`);
- paginação profunda limitada (~2.500 resultados por busca);
- HTTP 206 é resposta normal de sucesso;
- endpoint não é contrato público formal da VTEX e pode mudar sem aviso.
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from farol_common import num, price_stats, similarity, strip_accents, norm, STOPWORDS

SEARCH_PATH = "/api/catalog_system/pub/products/search"

# Lojas de exemplo confirmadas como VTEX via `verificar` em 2026-07. São apenas
# demonstração: configure via --lojas domínios VTEX do ramo do objeto licitado
# e valide com `vtex_catalog.py verificar` antes de usar em produção.
DEFAULT_STORES = ["www.epocacosmeticos.com.br", "www.cobasi.com.br", "www.brastemp.com.br"]

# Faixas de confiança do match descrição DFD x produto de varejo (Jaccard).
SIM_ALTA = 0.45
SIM_MEDIA = 0.25

MAX_WINDOW = 49  # a API rejeita janelas com mais de 50 itens


def eprint(*args: Any) -> None:
    print(*args, file=sys.stderr)


def agora_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def build_query(descricao: Any, max_termos: int = 5) -> str:
    """Reduz a descrição técnica do DFD a termos de busca de varejo.

    Mantém a ordem original das palavras (relevante para o ranking da busca
    VTEX), removendo acentos, stopwords e termos curtos.
    """
    words = strip_accents(norm(descricao)).split(" ")
    seen: List[str] = []
    for w in words:
        w = "".join(ch for ch in w if ch.isalnum())
        if len(w) >= 3 and w not in STOPWORDS and w not in seen:
            seen.append(w)
        if len(seen) >= max_termos:
            break
    return " ".join(seen).lower()


def fetch_store(loja: str, query: str, *, offset: int = 0, quantidade: int = 20,
                cache_dir: str | Path | None = None, retries: int = 3) -> tuple[List[Dict[str, Any]], str]:
    """Busca produtos em uma loja VTEX; retorna (payload, url_consulta).

    O domínio da loja entra no `path` (com base_url "https://") para que a
    chave de cache de `fetch_json` diferencie lojas distintas.
    """
    from farol_common import fetch_json

    quantidade = max(1, min(int(quantidade), MAX_WINDOW + 1))
    params = {"ft": query, "_from": offset, "_to": offset + quantidade - 1}
    path = f"{loja}{SEARCH_PATH}"
    from urllib.parse import urlencode
    url = "https://" + path + "?" + urlencode(params)
    payload = fetch_json(path, params, base_url="https://", cache_dir=cache_dir, retries=retries)
    if not isinstance(payload, list):
        return [], url
    return payload, url


def extract_offers(products: List[Dict[str, Any]], *, loja: str, url_consulta: str,
                   consultado_em: str, descricao_ref: Any = None) -> List[Dict[str, Any]]:
    """Achata o JSON hierárquico da VTEX (produto → SKU → vendedor → oferta).

    Cada linha de saída é uma cotação individual com os metadados de
    rastreabilidade exigidos pela IN 65/2021 (loja, URL, data/hora).
    """
    offers: List[Dict[str, Any]] = []
    for prod in products or []:
        if not isinstance(prod, dict):
            continue
        nome = prod.get("productName") or ""
        link = prod.get("link") or ""
        sim = round(similarity(descricao_ref, nome), 3) if descricao_ref else None
        for item in prod.get("items") or []:
            if not isinstance(item, dict):
                continue
            for seller in item.get("sellers") or []:
                offer = (seller or {}).get("commertialOffer") or {}
                preco = num(offer.get("Price"))
                disponivel = bool(offer.get("IsAvailable", (offer.get("AvailableQuantity") or 0) > 0))
                if preco is None or preco <= 0:
                    continue
                offers.append({
                    "loja": loja,
                    "produto": nome,
                    "sku": item.get("itemId"),
                    "nome_sku": item.get("name"),
                    "vendedor": (seller or {}).get("sellerName"),
                    "preco": preco,
                    "disponivel": disponivel,
                    "estoque": offer.get("AvailableQuantity"),
                    "link_produto": link,
                    "similaridade": sim,
                    "url_consulta": url_consulta,
                    "consultado_em": consultado_em,
                })
    return offers


def confidence_label(sim: float | None) -> str:
    if sim is None or sim <= 0:
        return "sem_match"
    if sim >= SIM_ALTA:
        return "alta"
    if sim >= SIM_MEDIA:
        return "media"
    return "baixa"


def summarize_item(descricao: Any, offers: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Resume as cotações de um item do DFD com estatísticas e confiança.

    As estatísticas de preço consideram apenas ofertas disponíveis cuja
    similaridade alcança pelo menos a faixa "media"; abaixo disso a cotação
    fica registrada como evidência, mas não entra na comparação.
    """
    validas = [o for o in offers if o.get("disponivel") and (o.get("preco") or 0) > 0]
    melhor = max((o.get("similaridade") or 0.0 for o in validas), default=0.0)
    confianca = confidence_label(melhor)
    comparaveis = [o for o in validas if (o.get("similaridade") or 0.0) >= SIM_MEDIA]
    resumo: Dict[str, Any] = {
        "consulta": build_query(descricao),
        "ofertas": len(validas),
        "comparaveis": len(comparaveis),
        "melhor_similaridade": round(melhor, 3),
        "confianca": confianca,
    }
    precos = [o["preco"] for o in comparaveis]
    resumo.update(price_stats(precos))
    if comparaveis:
        top = max(comparaveis, key=lambda o: o.get("similaridade") or 0.0)
        resumo.update({
            "melhor_produto": top.get("produto"),
            "melhor_loja": top.get("loja"),
            "melhor_preco_match": top.get("preco"),
            "melhor_link": top.get("link_produto"),
        })
    return resumo


def classify_varejo(estimado: float | None, mediana: float | None, confianca: str) -> str:
    """Avaliação do preço estimado x mediana de varejo, condicionada ao match.

    Alerta automático só nasce de match de confiança alta; nas demais faixas
    a comparação é rebaixada a indicativa, exigindo revisão humana — regra
    central para o varejo não contaminar a credibilidade dos achados.
    """
    if estimado is None or mediana is None or mediana <= 0:
        return "Sem comparação de varejo suficiente"
    ratio = estimado / mediana
    if confianca != "alta":
        return (f"Indicativo apenas ({ratio:.2f}x da mediana de varejo; confiança {confianca}) — "
                "revisar equivalência do produto antes de usar")
    if ratio >= 1.5:
        return f"Acima da mediana de varejo VTEX ({ratio:.2f}x); revisar estimativa (fonte complementar IN 65/2021)"
    if ratio <= 0.55:
        return f"Abaixo da mediana de varejo VTEX ({ratio:.2f}x); verificar subestimativa ou especificação (fonte complementar IN 65/2021)"
    return f"Compatível com varejo VTEX ({ratio:.2f}x; fonte complementar IN 65/2021)"


def pesquisar_item(descricao: Any, lojas: List[str], *, quantidade: int = 20,
                   sleep: float = 0.2, cache_dir: str | Path | None = None) -> tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """Pesquisa um item em todas as lojas e retorna (resumo, evidências)."""
    query = build_query(descricao)
    evidencias: List[Dict[str, Any]] = []
    if not query:
        return {"consulta": "", "ofertas": 0, "comparaveis": 0, "confianca": "sem_match"}, evidencias
    for loja in lojas:
        ts = agora_iso()
        try:
            products, url = fetch_store(loja, query, quantidade=quantidade, cache_dir=cache_dir)
        except Exception as exc:
            eprint(f"  aviso: falha em {loja}: {str(exc)[:120]}")
            continue
        evidencias.extend(extract_offers(products, loja=loja, url_consulta=url,
                                         consultado_em=ts, descricao_ref=descricao))
        if sleep:
            time.sleep(sleep)
    return summarize_item(descricao, evidencias), evidencias


def read_xlsx_items(path: str) -> List[Dict[str, Any]]:
    """Extrai (linha, código, descrição) da planilha DFD."""
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise SystemExit("Para ler XLSX instale openpyxl: pip install -r requirements.txt") from exc
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    code_col = None
    desc_col = None
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for r in range(1, min(max_row, 40) + 1):
        vals = [str(ws.cell(r, c).value or "").strip().upper() for c in range(1, min(max_col, 60) + 1)]
        for i, v in enumerate(vals, start=1):
            if "CÓDIGO" in v or "CODIGO" in v:
                header_row = r
                code_col = i
            if "DESCRI" in v:
                desc_col = i
        if header_row and desc_col:
            break
    if not desc_col:
        raise SystemExit("Não encontrei coluna DESCRIÇÃO na planilha.")
    items: List[Dict[str, Any]] = []
    for r in range((header_row or 1) + 1, max_row + 1):
        desc = ws.cell(r, desc_col).value
        if not desc or not str(desc).strip():
            continue
        codigo = None
        if code_col:
            try:
                codigo = int(float(ws.cell(r, code_col).value))
            except Exception:
                codigo = None
        items.append({"linha": r, "codigo": codigo, "descricao": str(desc).strip()})
    return items


def write_csv(rows: List[Dict[str, Any]], path: Path) -> None:
    import csv
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    fields: List[str] = []
    for r in rows:
        for k in r.keys():
            if k not in fields:
                fields.append(k)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def print_json(obj: Any) -> None:
    print(json.dumps(obj, ensure_ascii=False, indent=2, default=str))


def pesquisar_planilha(xlsx: str, lojas: List[str], outdir: Path, *, quantidade: int = 20,
                       sleep: float = 0.2, max_itens: int | None = None,
                       cache_dir: str | Path | None = None) -> Dict[str, Any]:
    """Pesquisa todos os itens da planilha DFD; grava resumo e evidências."""
    items = read_xlsx_items(xlsx)
    if max_itens:
        items = items[:max_itens]
    outdir.mkdir(parents=True, exist_ok=True)
    resumos: List[Dict[str, Any]] = []
    evidencias: List[Dict[str, Any]] = []
    for i, item in enumerate(items, start=1):
        eprint(f"[{i}/{len(items)}] varejo VTEX: {item['descricao'][:60]}...")
        resumo, evid = pesquisar_item(item["descricao"], lojas, quantidade=quantidade,
                                      sleep=sleep, cache_dir=cache_dir)
        resumo = {"linha": item["linha"], "codigo": item["codigo"], "descricao": item["descricao"], **resumo}
        for e in evid:
            e["linha_dfd"] = item["linha"]
            e["codigo_dfd"] = item["codigo"]
        resumos.append(resumo)
        evidencias.extend(evid)
    resumo_json = outdir / "resumo_varejo_por_item.json"
    resumo_json.write_text(json.dumps(resumos, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    write_csv(resumos, outdir / "resumo_varejo_por_item.csv")
    evid_json = outdir / "evidencias_varejo.json"
    evid_json.write_text(json.dumps(evidencias, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    write_csv(evidencias, outdir / "evidencias_varejo.csv")
    return {
        "planilha": xlsx,
        "lojas": lojas,
        "itens_pesquisados": len(items),
        "itens_com_oferta": sum(1 for r in resumos if r.get("ofertas")),
        "evidencias": len(evidencias),
        "resumo_json": str(resumo_json),
        "evidencias_json": str(evid_json),
    }


def parse_lojas(arg: str | None) -> List[str]:
    if not arg:
        return list(DEFAULT_STORES)
    return [x.strip().replace("https://", "").rstrip("/") for x in arg.split(",") if x.strip()]


def cmd_buscar(args: argparse.Namespace) -> None:
    resumo, evid = pesquisar_item(args.descricao, parse_lojas(args.lojas),
                                  quantidade=args.quantidade, sleep=args.sleep,
                                  cache_dir=args.cache)
    evid.sort(key=lambda o: -(o.get("similaridade") or 0.0))
    print_json({"resumo": resumo, "ofertas": evid[: args.limit]})
    if args.csv:
        write_csv(evid, Path(args.csv))
        eprint(f"CSV salvo em {args.csv}")


def cmd_planilha(args: argparse.Namespace) -> None:
    out = pesquisar_planilha(args.xlsx, parse_lojas(args.lojas), Path(args.out),
                             quantidade=args.quantidade, sleep=args.sleep,
                             max_itens=args.max_itens, cache_dir=args.cache)
    print_json(out)


def cmd_verificar(args: argparse.Namespace) -> None:
    """Confirma quais domínios respondem como loja VTEX (JSON array na busca)."""
    status = []
    for loja in parse_lojas(args.lojas):
        try:
            products, url = fetch_store(loja, "caneta", quantidade=1, retries=1)
            ok = isinstance(products, list)
            status.append({"loja": loja, "vtex": ok, "produtos_teste": len(products), "url": url})
        except Exception as exc:
            status.append({"loja": loja, "vtex": False, "erro": str(exc)[:120]})
    print_json(status)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="vtex-catalog",
                                description="Cotações de varejo via API pública de catálogo VTEX (fonte complementar IN 65/2021).")
    p.add_argument("--cache", help="diretório de cache local das respostas (evita repetir consultas idênticas)")
    sub = p.add_subparsers(dest="cmd", required=True)

    s = sub.add_parser("buscar", help="pesquisar um item por descrição livre nas lojas configuradas")
    s.add_argument("--descricao", required=True)
    s.add_argument("--lojas", help="domínios VTEX separados por vírgula (padrão: %s)" % ", ".join(DEFAULT_STORES))
    s.add_argument("--quantidade", type=int, default=20, help="ofertas por loja (máx. 50)")
    s.add_argument("--sleep", type=float, default=0.2)
    s.add_argument("--limit", type=int, default=10, help="ofertas exibidas no JSON")
    s.add_argument("--csv", help="salvar todas as cotações em CSV")
    s.set_defaults(func=cmd_buscar)

    s = sub.add_parser("planilha-precos", help="pesquisar todos os itens de uma planilha DFD e gerar resumo + evidências")
    s.add_argument("xlsx")
    s.add_argument("--out", default="output/vtex-varejo")
    s.add_argument("--lojas")
    s.add_argument("--quantidade", type=int, default=20)
    s.add_argument("--sleep", type=float, default=0.2)
    s.add_argument("--max-itens", type=int, help="limita itens pesquisados para teste/demonstração")
    s.set_defaults(func=cmd_planilha)

    s = sub.add_parser("verificar", help="testar se os domínios informados respondem como loja VTEX")
    s.add_argument("--lojas")
    s.set_defaults(func=cmd_verificar)
    return p


def main(argv: List[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    args.func(args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
