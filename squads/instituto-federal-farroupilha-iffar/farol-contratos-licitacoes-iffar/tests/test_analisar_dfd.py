import openpyxl
from openpyxl.styles import PatternFill

from analisar_dfd import (
    compras_price_findings,
    description_findings,
    is_yellow,
    outlier_findings,
    price_findings,
    risk_rank,
    total_consistency_findings,
)


def has_finding(findings, sev=None, tipo=None):
    return any((sev is None or f[0] == sev) and (tipo is None or f[1] == tipo) for f in findings)


def test_short_description_is_high_risk():
    fs = description_findings("COPO PLÁSTICO", "UNIDADE")
    assert has_finding(fs, sev="ALTO", tipo="DESCRIÇÃO")


def test_restrictive_term_flagged():
    fs = description_findings("PANELA EM ALUMÍNIO 7 LITROS MARCA TRAMONTINA COM VÁLVULA DE SEGURANÇA E CABO", "UNIDADE")
    assert any("MARCA" in f[2] for f in fs)


def test_restrictive_term_allows_brand_used_only_as_reference():
    # Regressão: marca citada apenas como referência de qualidade, com similar/
    # equivalente admitido, é uso permitido pelo art. 41 da Lei nº 14.133/2021
    # e não deveria ser apontada como termo restritivo.
    fs = description_findings(
        "PANELA MARCA DE REFERÊNCIA TRAMONTINA 7 LITROS EM ALUMÍNIO ADMITIDA SIMILAR DE QUALIDADE EQUIVALENTE",
        "UNIDADE",
    )
    assert not any("MARCA" in f[2] for f in fs)


def test_unit_package_mismatch():
    fs = description_findings("PAPEL TOALHA INTERFOLHADO PACOTE COM 1000 FOLHAS 2 DOBRAS 100% CELULOSE", "UNIDADE")
    assert has_finding(fs, tipo="UNIDADE")
    fs2 = description_findings("LUVA DE PROCEDIMENTO EM LÁTEX TAMANHO M AMBIDESTRA DESCARTÁVEL COM PÓ", "CAIXA")
    assert has_finding(fs2, tipo="UNIDADE")


def test_price_findings():
    assert has_finding(price_findings(None), sev="MÉDIO", tipo="PREÇO")
    assert has_finding(price_findings(0), sev="ALTO", tipo="PREÇO")
    assert price_findings(9.90) == []


def test_total_consistency():
    fs = total_consistency_findings(10.0, 100, 5000.0)
    assert has_finding(fs, sev="MÉDIO", tipo="PREÇO")
    assert total_consistency_findings(10.0, 100, 1000.0) == []
    assert total_consistency_findings(None, 100, 5000.0) == []


def test_outlier_high_quantity_detected():
    vals = [("A", 20), ("B", 18), ("C", 500), ("D", 22), ("E", 19)]
    fs = outlier_findings(vals)
    assert any(f[0] == "ALTO" and "C" in f[2] for f in fs)


def test_outlier_zero_demand_with_wide_adoption():
    vals = [(f"C{i}", 30) for i in range(7)] + [("Z", 0)]
    fs = outlier_findings(vals)
    assert any(f[0] == "BAIXO" and "Z" in f[2] for f in fs)


def test_outlier_needs_minimum_sample():
    assert outlier_findings([("A", 10), ("B", 500)]) == []


def test_compras_price_findings():
    stats = {"registros": 10, "mediana": 10.0}
    assert has_finding(compras_price_findings(30.0, stats), sev="ALTO", tipo="PREÇO EXTERNO")
    assert has_finding(compras_price_findings(2.0, stats), sev="MÉDIO", tipo="PREÇO EXTERNO")
    assert compras_price_findings(11.0, stats) == []
    assert has_finding(compras_price_findings(11.0, {"erro": "timeout"}), tipo="COMPRAS.GOV")


def test_risk_rank():
    assert risk_rank(["BAIXO", "ALTO", "MÉDIO"]) == "ALTO"
    assert risk_rank(["BAIXO", "MÉDIO"]) == "MÉDIO"
    assert risk_rank([]) == "OK"


def _cell_with_fill(rgb):
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws.cell(1, 1, "x")
    cell.fill = PatternFill("solid", fgColor=rgb)
    return cell


def test_is_yellow_exact_and_heuristic():
    assert is_yellow(_cell_with_fill("FFFFFF00"))
    assert is_yellow(_cell_with_fill("FFFFE000"))  # amarelo com tint
    assert not is_yellow(_cell_with_fill("FF0000FF"))
    assert not is_yellow(_cell_with_fill("FFFFFFFF"))
