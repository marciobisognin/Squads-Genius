import csv
import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

from analisar_dfd import analyze, detect_columns, find_header
from gerar_dfd_exemplo import build

ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture(scope="module")
def fixture_xlsx(tmp_path_factory):
    return build(tmp_path_factory.mktemp("dfd") / "dfd_exemplo.xlsx")


def test_header_and_columns_detection(fixture_xlsx):
    wb = openpyxl.load_workbook(fixture_xlsx)
    ws = wb[wb.sheetnames[0]]
    header = find_header(ws)
    assert header == 3
    cols, qcols = detect_columns(ws, header)
    assert {"codigo", "descricao", "unidade", "preco", "valor_total"} <= set(cols)
    assert len(qcols) == 5
    assert qcols[0][1] == "Campus Alegrete"


def test_full_offline_analysis(fixture_xlsx, tmp_path):
    out = tmp_path / "auditoria"
    summary = analyze(fixture_xlsx, out)
    assert summary["items_analisados"] == 12
    assert len(summary["campi_detectados"]) == 5
    for key in ["planilha", "csv", "relatorio", "dashboard"]:
        assert Path(summary["outputs"][key]).exists()
    tipos = summary["tipos"]
    assert {"DESCRIÇÃO", "UNIDADE", "PREÇO", "OUTLIER"} <= set(tipos)
    assert summary["riscos"]["ALTO"] >= 2
    assert summary["valor_estimado_por_risco"]["ALTO"] > 0
    with Path(summary["outputs"]["csv"]).open(encoding="utf-8-sig") as f:
        achados = list(csv.DictReader(f))
    assert achados, "CSV de achados não pode estar vazio na fixture com problemas plantados"
    assert any("Valor total informado" in a["achado"] for a in achados), "divergência de valor total deve ser detectada"
    wb = openpyxl.load_workbook(summary["outputs"]["planilha"])
    ws = wb[wb.sheetnames[0]]
    headers = [str(ws.cell(3, c).value or "") for c in range(1, ws.max_column + 1)]
    assert "Valor Total Estimado (R$)" in headers
    assert "Nível de Risco" in headers


def test_saneamento_flow(fixture_xlsx, tmp_path):
    out = tmp_path / "auditoria"
    summary = analyze(fixture_xlsx, out)
    san_dir = tmp_path / "saneamento"
    script = ROOT / "scripts" / "painel_saneamento.py"
    subprocess.run([sys.executable, str(script), "gerar", summary["outputs"]["csv"], "--out", str(san_dir)], check=True, capture_output=True)
    san_csv = san_dir / "saneamento.csv"
    assert san_csv.exists()
    subprocess.run([sys.executable, str(script), "atualizar", str(san_csv), "--id", "1", "--status", "CORRIGIDO", "--obs", "teste"], check=True, capture_output=True)
    subprocess.run([sys.executable, str(script), "painel", str(san_csv), "--out", str(san_dir / "painel.html")], check=True, capture_output=True)
    assert (san_dir / "painel.html").exists()
    with san_csv.open(encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["status"] == "CORRIGIDO"


def test_historico_flow(fixture_xlsx, tmp_path):
    out = tmp_path / "auditoria"
    analyze(fixture_xlsx, out)
    hist = tmp_path / "historico"
    script = ROOT / "scripts" / "historico_farol.py"
    for ciclo in ["2025-1", "2026-1"]:
        subprocess.run([sys.executable, str(script), "registrar", str(out), "--ciclo", ciclo, "--historico", str(hist)], check=True, capture_output=True)
    proc = subprocess.run([sys.executable, str(script), "comparar", "--historico", str(hist), "--out", str(hist / "rel.md")], check=True, capture_output=True, text=True)
    result = json.loads(proc.stdout)
    assert result["ciclos"] == ["2025-1", "2026-1"]
    assert result["recorrencias"] > 0, "mesma planilha em dois ciclos deve gerar recorrências"
    assert (hist / "rel.md").exists()


def test_base_conhecimento_flow(fixture_xlsx, tmp_path):
    script = ROOT / "scripts" / "base_conhecimento.py"
    base = tmp_path / "kb.json"
    subprocess.run([
        sys.executable, str(script), "--base", str(base), "adicionar",
        "--codigo", "245123",
        "--descricao", "COPO PLÁSTICO DESCARTÁVEL 200 ML, POLIPROPILENO TRANSPARENTE, PACOTE COM 100 UNIDADES",
        "--unidade", "PACOTE",
    ], check=True, capture_output=True)
    out_csv = tmp_path / "verificacao.csv"
    proc = subprocess.run([
        sys.executable, str(script), "--base", str(base), "verificar", str(fixture_xlsx), "--out", str(out_csv),
    ], check=True, capture_output=True, text=True)
    result = json.loads(proc.stdout)
    assert result["itens_na_base"] == 1
    assert result["divergentes"] == 1, "descrição curta da planilha diverge da aprovada na base"


def test_previsao_quantitativos(fixture_xlsx, tmp_path):
    script = ROOT / "scripts" / "previsao_quantitativos.py"
    proc = subprocess.run([
        sys.executable, str(script), str(fixture_xlsx), str(fixture_xlsx), "--out", str(tmp_path / "previsao"),
    ], check=True, capture_output=True, text=True)
    result = json.loads(proc.stdout)
    assert result["pares_item_campus_comparados"] > 0
    assert Path(result["csv"]).exists()
