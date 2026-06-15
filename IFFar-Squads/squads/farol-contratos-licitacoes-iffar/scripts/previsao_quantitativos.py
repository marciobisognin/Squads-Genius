#!/usr/bin/env python3
"""Previsão simples de quantitativos por campus para o squad Farol IFFar.

Recebe planilhas DFD de ciclos anteriores e a planilha do ciclo atual.
Para cada par (código de item, campus), calcula a referência histórica
(mediana dos ciclos anteriores) e compara com o quantitativo atual,
sinalizando desvios relevantes.

Uso:
    python scripts/previsao_quantitativos.py dfd_2024.xlsx dfd_2025.xlsx dfd_2026.xlsx --out output/previsao
    # a última planilha é tratada como o ciclo atual; as demais, como histórico
"""
from __future__ import annotations

import argparse
import csv
import json
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl

from analisar_dfd import detect_columns, find_header
from farol_common import num

DESVIO_ALTO = 2.0
DESVIO_BAIXO = 0.5


def read_quantities(path: str) -> Dict[Tuple[str, str], float]:
    """Extrai {(codigo, campus): quantidade} de uma planilha DFD."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = find_header(ws)
    cols, qcols = detect_columns(ws, header)
    data: Dict[Tuple[str, str], float] = {}
    for r in range(header + 1, ws.max_row + 1):
        desc = ws.cell(r, cols.get("descricao", 3)).value
        if not desc or not str(desc).strip():
            continue
        codigo = ws.cell(r, cols.get("codigo", 2)).value
        try:
            codigo = str(int(float(codigo)))
        except Exception:
            continue
        for c, campus in qcols:
            v = num(ws.cell(r, c).value)
            if v is not None:
                data[(codigo, str(campus))] = v
    return data


def main() -> int:
    ap = argparse.ArgumentParser(description="Compara quantitativos do ciclo atual com a referência histórica por item/campus.")
    ap.add_argument("planilhas", nargs="+", help="planilhas DFD em ordem cronológica; a última é o ciclo atual")
    ap.add_argument("--out", default="output/previsao")
    args = ap.parse_args()
    if len(args.planilhas) < 2:
        raise SystemExit("Informe pelo menos duas planilhas: histórico... + ciclo atual.")
    historico = [read_quantities(p) for p in args.planilhas[:-1]]
    atual = read_quantities(args.planilhas[-1])
    series: Dict[Tuple[str, str], List[float]] = defaultdict(list)
    for ciclo in historico:
        for key, v in ciclo.items():
            series[key].append(v)
    rows: List[Dict[str, Any]] = []
    desvios = 0
    for (codigo, campus), atual_v in sorted(atual.items()):
        hist = series.get((codigo, campus))
        if not hist:
            continue
        referencia = statistics.median(hist)
        if referencia > 0 and atual_v > 0:
            razao = atual_v / referencia
            if razao >= DESVIO_ALTO:
                situacao = f"ACIMA DO HISTÓRICO ({razao:.1f}x); confirmar demanda com o campus"
                desvios += 1
            elif razao <= DESVIO_BAIXO:
                situacao = f"ABAIXO DO HISTÓRICO ({razao:.1f}x); confirmar se a redução é intencional"
                desvios += 1
            else:
                situacao = "Dentro da faixa histórica"
        elif referencia > 0 and atual_v == 0:
            situacao = "Sem demanda no ciclo atual, mas com histórico de consumo; confirmar"
            desvios += 1
            razao = 0.0
        else:
            situacao = "Sem referência comparável"
            razao = None
        rows.append({
            "codigo": codigo,
            "campus": campus,
            "ciclos_historicos": len(hist),
            "referencia_historica": referencia,
            "quantidade_atual": atual_v,
            "razao": round(razao, 2) if razao is not None else "",
            "situacao": situacao,
        })
    outdir = Path(args.out)
    outdir.mkdir(parents=True, exist_ok=True)
    csv_path = outdir / "previsao_quantitativos.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["codigo", "campus", "ciclos_historicos", "referencia_historica", "quantidade_atual", "razao", "situacao"])
        w.writeheader()
        w.writerows(rows)
    summary = {
        "planilhas_historico": args.planilhas[:-1],
        "planilha_atual": args.planilhas[-1],
        "pares_item_campus_comparados": len(rows),
        "desvios_sinalizados": desvios,
        "csv": str(csv_path),
    }
    (outdir / "previsao_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
