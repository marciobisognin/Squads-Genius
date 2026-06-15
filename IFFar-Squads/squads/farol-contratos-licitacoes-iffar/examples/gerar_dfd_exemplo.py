#!/usr/bin/env python3
"""Gera uma planilha DFD fictícia no formato multicampi do IFFar.

Os dados são sintéticos e contêm problemas plantados de propósito (descrição
curta, termo restritivo, unidade incompatível, preço ausente/zero, outlier de
quantitativo e valor total divergente) para demonstração e testes do squad.

Uso:
    python examples/gerar_dfd_exemplo.py [--out examples/dfd_exemplo.xlsx]
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

CAMPI = ["Campus Alegrete", "Campus Jaguari", "Campus Santo Augusto", "Campus São Borja", "Reitoria"]

# (codigo, descricao, unidade, preco, [quantidades por campus], valor_total_override)
ITENS = [
    (437939, "CANETA ESFEROGRÁFICA AZUL, ESCRITA MÉDIA 1.0 MM, CORPO EM RESINA TERMOPLÁSTICA TRANSPARENTE, CARGA REMOVÍVEL", "UNIDADE", 1.20, [100, 120, 90, 110, 95], None),
    (245123, "COPO PLÁSTICO", "UNIDADE", 0.15, [200, 180, 220, 190, 210], None),
    (358441, "PANELA DE PRESSÃO EM ALUMÍNIO POLIDO, CAPACIDADE 7 LITROS, MARCA TRAMONTINA OU SIMILAR, VÁLVULA DE SEGURANÇA", "UNIDADE", 145.00, [4, 5, 3, 4, 6], None),
    (412877, "PAPEL TOALHA INTERFOLHADO, PACOTE COM 1000 FOLHAS, 2 DOBRAS, 100% CELULOSE VIRGEM, 20X21 CM", "UNIDADE", 12.50, [40, 35, 50, 45, 38], None),
    (398210, "LUVA DE PROCEDIMENTO EM LÁTEX, TAMANHO M, AMBIDESTRA, COM PÓ BIOABSORVÍVEL, DESCARTÁVEL", "CAIXA", 28.90, [25, 30, 20, 28, 22], None),
    (501998, "DETERGENTE LÍQUIDO NEUTRO PARA LOUÇAS, FRASCO 500 ML, BIODEGRADÁVEL, TESTADO DERMATOLOGICAMENTE", "FRASCO", None, [60, 55, 70, 65, 58], None),
    (476201, "ÁLCOOL ETÍLICO HIDRATADO 70%, FRASCO 1 LITRO, USO GERAL EM LIMPEZA E DESINFECÇÃO DE SUPERFÍCIES", "FRASCO", 0.0, [80, 75, 90, 85, 78], None),
    (333456, "GRAMPEADOR DE MESA METÁLICO PARA 25 FOLHAS, GRAMPO 26/6, BASE EMBORRACHADA ANTIDERRAPANTE", "UNIDADE", 22.40, [20, 18, 500, 22, 19], None),
    (287654, "CADERNO UNIVERSITÁRIO CAPA DURA, 96 FOLHAS PAUTADAS, ESPIRAL METÁLICO, FORMATO 200X275 MM", "UNIDADE", 10.00, [20, 20, 20, 20, 20], 5000.00),
    (419302, "CADO DE AÇO INXIDÁVEL PARA FACA DE COZINHA PROFISSIONAL, LÂMINA 8 POLEGADAS, EMPUNHADURA ERGONÔMICA", "UNIDADE", 35.00, [6, 8, 5, 7, 6], None),
    (None, "TESOURA MULTIUSO EM AÇO INOXIDÁVEL, 21 CM, CABO EM POLIPROPILENO, PONTA ARREDONDADA, USO ESCOLAR", "UNIDADE", 8.75, [30, 25, 35, 28, 32], None),
    (529871, "GARRAFA TÉRMICA EM AÇO INOXIDÁVEL, CAPACIDADE 1,8 LITROS, AMPOLA INTERNA EM VIDRO, ALÇA TRANSPORTE", "UNIDADE", 89.90, [10, 12, 8, 11, 9], None),
]

YELLOW = PatternFill("solid", fgColor="FFFFFF00")
HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)


def build(out_path: str | Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DFD"
    ws.cell(1, 1, "DFD ND 339030.99 — MATERIAIS DE EXEMPLO (DADOS FICTÍCIOS PARA DEMONSTRAÇÃO)")
    ws.cell(1, 1).font = Font(bold=True)
    fixed = ["Nº", "CÓDIGO", "DESCRIÇÃO DO ITEM", "UNIDADE DE FORNECIMENTO", "VALOR ESTIMADO NA ÚLTIMA COMPRA (R$)"]
    header_row = 3
    for i, h in enumerate(fixed, start=1):
        cell = ws.cell(header_row, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    qty_start = len(fixed) + 1
    for j, campus in enumerate(CAMPI):
        col = qty_start + j
        ws.cell(2, col, campus).font = Font(bold=True)
        cell = ws.cell(header_row, col, "QUANTIDADE ESTIMADA")
        cell.fill = YELLOW
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    total_col = qty_start + len(CAMPI)
    cell = ws.cell(header_row, total_col, "VALOR TOTAL ESTIMADO (R$)")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    for n, (codigo, desc, unidade, preco, qts, total_override) in enumerate(ITENS, start=1):
        r = header_row + n
        ws.cell(r, 1, n)
        ws.cell(r, 2, codigo)
        ws.cell(r, 3, desc)
        ws.cell(r, 4, unidade)
        ws.cell(r, 5, preco)
        for j, q in enumerate(qts):
            c = ws.cell(r, qty_start + j, q)
            c.fill = YELLOW
        total = total_override if total_override is not None else (round(preco * sum(qts), 2) if preco else None)
        ws.cell(r, total_col, total)
    widths = {1: 5, 2: 10, 3: 60, 4: 16, 5: 16, total_col: 18}
    for col, w in widths.items():
        ws.column_dimensions[ws.cell(header_row, col).column_letter].width = w
    for j in range(len(CAMPI)):
        ws.column_dimensions[ws.cell(header_row, qty_start + j).column_letter].width = 14
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    ap = argparse.ArgumentParser(description="Gera planilha DFD fictícia para demonstração do squad.")
    ap.add_argument("--out", default=str(Path(__file__).resolve().parent / "dfd_exemplo.xlsx"))
    args = ap.parse_args()
    path = build(args.out)
    print(f"Planilha de exemplo gerada em {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
