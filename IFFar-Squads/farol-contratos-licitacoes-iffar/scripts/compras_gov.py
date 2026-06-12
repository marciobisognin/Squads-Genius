#!/usr/bin/env python3
"""compras-gov: CLI leve para Dados Abertos Compras.gov.br.

Integra pesquisa de itens, preços praticados, atas de registro de preços,
editais/licitações legadas e contratações PNCP/Lei 14.133.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import statistics
import sys
import time
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

from farol_common import similarity

BASE_URL = "https://dadosabertos.compras.gov.br"
USER_AGENT = "compras-gov-cli/1.3 (+farol-contratos-licitacoes-iffar; dadosabertos.compras.gov.br)"

# Diretório de cache das respostas da API; definido via --cache no comando principal.
CACHE_DIR: Path | None = None

ENDPOINTS = {
    "pdm-material": "/modulo-material/3_consultarPdmMaterial",
    "material-item": "/modulo-material/4_consultarItemMaterial",
    "precos-material": "/modulo-pesquisa-preco/1_consultarMaterial",
    "precos-material-detalhe": "/modulo-pesquisa-preco/2_consultarMaterialDetalhe",
    "precos-servico": "/modulo-pesquisa-preco/3_consultarServico",
    "precos-servico-detalhe": "/modulo-pesquisa-preco/4_consultarServicoDetalhe",
    "licitacoes-legado": "/modulo-legado/1_consultarLicitacao",
    "itens-licitacao-legado": "/modulo-legado/2_consultarItemLicitacao",
    "contratacoes-pncp": "/modulo-contratacoes/1_consultarContratacoes_PNCP_14133",
    "itens-contratacoes-pncp": "/modulo-contratacoes/2_consultarItensContratacoes_PNCP_14133",
    "resultados-pncp": "/modulo-contratacoes/3_consultarResultadoItensContratacoes_PNCP_14133",
    "atas-item": "/modulo-arp/2_consultarARPItem",
    "unidades-ata-item": "/modulo-arp/3_consultarUnidadesItem",
    "empenhos-saldo-item": "/modulo-arp/4_consultarEmpenhosSaldoItem",
    "adesoes-item": "/modulo-arp/5_consultarAdesoesItem",
}

DEFAULT_FIELDS = {
    "precos-material": ["dataCompra", "codigoItemCatalogo", "precoUnitario", "quantidade", "descricaoItem", "nomeUnidadeFornecimento", "nomeFornecedor", "codigoUasg", "nomeUasg", "estado", "municipio", "idCompra"],
    "atas-item": ["numeroAtaRegistroPreco", "codigoUnidadeGerenciadora", "numeroCompra", "anoCompra", "codigoItem", "valorUnitario", "quantidadeHomologadaItem", "descricaoItem", "nomeRazaoSocialFornecedor", "dataVigenciaInicial", "dataVigenciaFinal", "nomeUnidadeGerenciadora", "numeroControlePncpAta"],
    "licitacoes-legado": ["uasg", "modalidade", "numero_aviso", "objeto", "data_publicacao", "data_abertura_proposta", "situacao_aviso"],
    "contratacoes-pncp": ["numeroControlePNCP", "modalidadeNome", "objetoCompra", "dataPublicacaoPncp", "unidadeOrgaoNomeUnidade", "unidadeOrgaoUfSigla", "valorTotalEstimado"],
    "itens-contratacoes-pncp": ["numeroControlePNCP", "numeroItem", "codItemCatalogo", "descricao", "materialOuServico", "quantidade", "valorUnitarioEstimado", "valorTotal", "unidadeOrgaoNomeUnidade"],
    "resultados-pncp": ["numeroControlePNCP", "numeroItem", "niFornecedor", "nomeRazaoSocialFornecedor", "valorUnitarioHomologado", "valorTotalHomologado", "dataResultadoPncp"],
    "material-item": ["codigoItem", "descricaoItem", "codigoClasse", "nomeClasse", "codigoGrupo", "nomeGrupo", "statusItem"],
}


def eprint(*args: Any) -> None:
    print(*args, file=sys.stderr)


INTERNAL_ARGS = {"cmd", "paginas", "sleep", "format", "csv", "out_json", "limit", "func", "xlsx", "out", "max_itens", "cache", "tipo", "descricao"}


def clean_params(params: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k, v in params.items():
        if k in INTERNAL_ARGS or v is None or v == "":
            continue
        if isinstance(v, bool):
            out[k] = str(v).lower()
        else:
            out[k] = v
    return out


def fetch(path: str, params: Dict[str, Any], *, raw: bool = False, retries: int = 3, backoff: float = 1.5) -> Any:
    url = BASE_URL + path + "?" + urlencode(clean_params(params), doseq=True)
    cache_file = None
    if CACHE_DIR is not None and not raw:
        cache_file = CACHE_DIR / (hashlib.sha256(url.encode("utf-8")).hexdigest()[:32] + ".json")
        if cache_file.exists():
            try:
                return json.loads(cache_file.read_text(encoding="utf-8"))
            except Exception:
                pass
    req = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "application/json,text/csv,*/*"})
    text = ""
    ct = ""
    for attempt in range(retries + 1):
        try:
            with urlopen(req, timeout=90) as resp:
                data = resp.read()
                text = data.decode("utf-8", errors="replace")
                ct = resp.headers.get("content-type", "")
            break
        except HTTPError as ex:
            if attempt < retries and ex.code in (429, 500, 502, 503, 504):
                time.sleep(backoff * (2 ** attempt))
                continue
            body = ex.read().decode("utf-8", errors="replace")
            raise SystemExit(f"HTTP {ex.code} em {url}\n{body[:1200]}")
        except (URLError, TimeoutError, OSError) as ex:
            if attempt < retries:
                time.sleep(backoff * (2 ** attempt))
                continue
            raise SystemExit(f"Erro de rede em {url}: {ex}")
    if raw:
        return text
    if "json" in ct or text.strip().startswith(("{", "[")):
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            return {"raw": text, "url": url}
        if cache_file is not None:
            cache_file.parent.mkdir(parents=True, exist_ok=True)
            cache_file.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        return payload
    return {"raw": text, "url": url}


def extract_rows(payload: Any) -> List[Dict[str, Any]]:
    if isinstance(payload, dict):
        rows = payload.get("resultado") or payload.get("results") or payload.get("data") or []
        return rows if isinstance(rows, list) else []
    return payload if isinstance(payload, list) else []


def fetch_pages(path: str, params: Dict[str, Any], max_pages: int = 1, sleep: float = 0.0) -> Dict[str, Any]:
    params = dict(params)
    params.setdefault("pagina", 1)
    params.setdefault("tamanhoPagina", 10)
    rows: List[Dict[str, Any]] = []
    last: Dict[str, Any] = {}
    for page in range(int(params.get("pagina", 1)), int(params.get("pagina", 1)) + max_pages):
        params["pagina"] = page
        payload = fetch(path, params)
        if not isinstance(payload, dict):
            break
        last = payload
        batch = extract_rows(payload)
        rows.extend(batch)
        total_pages = int(payload.get("totalPaginas") or 0)
        if total_pages and page >= total_pages:
            break
        if not batch:
            break
        if sleep:
            time.sleep(sleep)
    last = dict(last)
    last["resultado"] = rows
    last["registrosRetornados"] = len(rows)
    return last


def print_json(obj: Any) -> None:
    print(json.dumps(obj, ensure_ascii=False, indent=2, default=str))


def flatten_value(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)
    return str(v)


def write_csv(rows: List[Dict[str, Any]], path: str | None, fields: List[str] | None = None) -> None:
    if fields is None:
        keys: List[str] = []
        for r in rows:
            for k in r.keys():
                if k not in keys:
                    keys.append(k)
        fields = keys
    out = open(path, "w", newline="", encoding="utf-8-sig") if path else sys.stdout
    close = path is not None
    try:
        w = csv.DictWriter(out, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: flatten_value(r.get(k)) for k in fields})
    finally:
        if close:
            out.close()


def table(rows: List[Dict[str, Any]], fields: List[str] | None = None, limit: int = 20) -> None:
    rows = rows[:limit]
    if not rows:
        print("Sem resultados.")
        return
    fields = fields or list(rows[0].keys())[:8]
    widths = []
    for f in fields:
        maxlen = max([len(str(f))] + [len(flatten_value(r.get(f))[:80]) for r in rows])
        widths.append(min(maxlen, 38))
    print(" | ".join(str(f)[:w].ljust(w) for f, w in zip(fields, widths)))
    print("-+-".join("-" * w for w in widths))
    for r in rows:
        vals = []
        for f, w in zip(fields, widths):
            val = flatten_value(r.get(f)).replace("\n", " ")[:w]
            vals.append(val.ljust(w))
        print(" | ".join(vals))


def output_payload(payload: Dict[str, Any], args: argparse.Namespace, endpoint_key: str | None = None) -> None:
    rows = extract_rows(payload)
    if args.out_json:
        Path(args.out_json).write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        eprint(f"JSON salvo em {args.out_json}")
    if args.csv:
        write_csv(rows, args.csv, DEFAULT_FIELDS.get(endpoint_key or ""))
        eprint(f"CSV salvo em {args.csv}")
    if args.format == "json":
        print_json(payload)
    elif args.format == "csv":
        write_csv(rows, None, DEFAULT_FIELDS.get(endpoint_key or ""))
    else:
        meta = {k: payload.get(k) for k in ["totalRegistros", "totalPaginas", "paginasRestantes", "registrosRetornados"] if isinstance(payload, dict) and k in payload}
        if meta:
            eprint("Resumo:", json.dumps(meta, ensure_ascii=False))
        table(rows, DEFAULT_FIELDS.get(endpoint_key or ""), args.limit)


def common_page_args(p: argparse.ArgumentParser) -> None:
    p.add_argument("--pagina", type=int, default=1)
    p.add_argument("--tamanho-pagina", type=int, default=10, dest="tamanhoPagina", help="API exige 10 a 500")
    p.add_argument("--paginas", type=int, default=1, help="quantas páginas buscar")
    p.add_argument("--sleep", type=float, default=0.0, help="pausa entre páginas")
    p.add_argument("--format", choices=["table", "json", "csv"], default="table")
    p.add_argument("--csv", help="salvar resultados em CSV")
    p.add_argument("--out-json", help="salvar payload completo em JSON")
    p.add_argument("--limit", type=int, default=20, help="linhas exibidas na tabela")


def cmd_endpoint(args: argparse.Namespace, key: str, params: Dict[str, Any]) -> None:
    payload = fetch_pages(ENDPOINTS[key], params, args.paginas, args.sleep)
    output_payload(payload, args, key)


def date_range_default(days: int = 365) -> Tuple[str, str]:
    end = date.today()
    start = end - timedelta(days=days)
    return start.isoformat(), end.isoformat()


def cmd_preco_resumo(args: argparse.Namespace) -> None:
    inicio, fim = args.inicio, args.fim
    payload = fetch_pages(ENDPOINTS["precos-material"], {
        "codigoItemCatalogo": args.codigo_item,
        "dataCompraInicio": inicio,
        "dataCompraFim": fim,
        "codigoUasg": args.uasg,
        "estado": args.uf,
        "pagina": args.pagina,
        "tamanhoPagina": args.tamanhoPagina,
    }, args.paginas, args.sleep)
    rows = extract_rows(payload)
    prices = [float(r.get("precoUnitario")) for r in rows if r.get("precoUnitario") not in (None, "") and float(r.get("precoUnitario")) > 0]
    resumo: Dict[str, Any] = {
        "codigoItemCatalogo": args.codigo_item,
        "periodo": {"inicio": inicio, "fim": fim},
        "registros": len(rows),
        "precos_validos": len(prices),
    }
    if prices:
        resumo.update({
            "min": min(prices),
            "max": max(prices),
            "media": sum(prices) / len(prices),
            "mediana": statistics.median(prices),
        })
        if len(prices) >= 4:
            q = statistics.quantiles(prices, n=4, method="inclusive")
            resumo.update({"q1": q[0], "q3": q[2], "iqr": q[2] - q[0]})
        resumo["amostra"] = rows[: min(len(rows), args.limit)]
    print_json(resumo)
    if args.csv:
        write_csv(rows, args.csv, DEFAULT_FIELDS["precos-material"])
        eprint(f"CSV salvo em {args.csv}")


def read_xlsx_codes(path: str) -> List[int]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise SystemExit("Para ler XLSX instale openpyxl: pip install -r requirements.txt") from exc
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    code_col = None
    desc_col = None
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for r in range(1, min(max_row, 30) + 1):
        vals = [str(ws.cell(r, c).value or "").strip().upper() for c in range(1, min(max_col, 20) + 1)]
        for i, v in enumerate(vals, start=1):
            if "CÓDIGO" in v or "CODIGO" in v:
                header_row = r
                code_col = i
            if "DESCRI" in v:
                desc_col = i
        if header_row and code_col:
            break
    if not code_col:
        raise SystemExit("Não encontrei coluna CÓDIGO/CODIGO na planilha.")
    codes = []
    for r in range((header_row or 1) + 1, max_row + 1):
        v = ws.cell(r, code_col).value
        try:
            code = int(float(v))
        except Exception:
            continue
        if code > 0 and code not in codes:
            codes.append(code)
    return codes


def load_pdm_catalog(cache_path: Path, max_paginas: int = 60, sleep: float = 0.1) -> List[Dict[str, Any]]:
    """Baixa o catálogo PDM completo uma única vez e mantém em cache local.

    O filtro textual `nomePdm` da API não funciona no servidor, então a busca
    por descrição precisa ser feita localmente sobre o catálogo inteiro.
    """
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))
    eprint("Baixando catálogo PDM completo (apenas na primeira execução; fica em cache local)...")
    rows: List[Dict[str, Any]] = []
    page = 1
    while page <= max_paginas:
        payload = fetch(ENDPOINTS["pdm-material"], {"pagina": page, "tamanhoPagina": 500})
        batch = extract_rows(payload)
        rows.extend(batch)
        total = int(payload.get("totalPaginas") or 0)
        eprint(f"  página {page}/{total or '?'} — {len(rows)} PDMs")
        if not batch or (total and page >= total):
            break
        page += 1
        if sleep:
            time.sleep(sleep)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")
    return rows


def cmd_sugerir_codigo(args: argparse.Namespace) -> None:
    """Sugere códigos CATMAT a partir de uma descrição livre, ranqueados por similaridade."""
    catalog = load_pdm_catalog(Path(args.catalogo_pdm), sleep=args.sleep)
    scored = sorted(
        ((similarity(args.descricao, p.get("nomePdm")), p) for p in catalog if p.get("statusPdm")),
        key=lambda x: -x[0],
    )
    top_pdms = [(s, p) for s, p in scored[: args.top_pdm] if s > 0]
    candidatos: List[Dict[str, Any]] = []
    for _, pdm in top_pdms:
        payload = fetch_pages(ENDPOINTS["material-item"], {
            "codigoPdm": pdm.get("codigoPdm"),
            "statusItem": True,
            "pagina": 1,
            "tamanhoPagina": 100,
        }, 1, args.sleep)
        for item in extract_rows(payload):
            candidatos.append({
                "codigoItem": item.get("codigoItem"),
                "descricaoItem": item.get("descricaoItem"),
                "nomePdm": pdm.get("nomePdm"),
                "nomeClasse": item.get("nomeClasse") or pdm.get("nomeClasse"),
                "similaridade": round(similarity(args.descricao, item.get("descricaoItem")), 3),
            })
    candidatos.sort(key=lambda c: -c["similaridade"])
    candidatos = candidatos[: args.limit]
    print_json({
        "descricao": args.descricao,
        "pdms_candidatos": [{"codigoPdm": p.get("codigoPdm"), "nomePdm": p.get("nomePdm"), "similaridade": round(s, 3)} for s, p in top_pdms],
        "candidatos": candidatos,
    })
    if args.csv:
        write_csv(candidatos, args.csv)
        eprint(f"CSV salvo em {args.csv}")


def cmd_planilha_pesquisar(args: argparse.Namespace) -> None:
    codes = read_xlsx_codes(args.xlsx)
    if args.max_itens:
        codes = codes[: args.max_itens]
    outdir = Path(args.out)
    outdir.mkdir(parents=True, exist_ok=True)
    endpoint = "precos-servico" if getattr(args, "tipo", "material") == "servico" else "precos-material"
    summary_rows: List[Dict[str, Any]] = []
    for i, code in enumerate(codes, start=1):
        eprint(f"[{i}/{len(codes)}] pesquisando código {code}...")
        payload = fetch_pages(ENDPOINTS[endpoint], {
            "codigoItemCatalogo": code,
            "dataCompraInicio": args.inicio,
            "dataCompraFim": args.fim,
            "pagina": 1,
            "tamanhoPagina": args.tamanhoPagina,
        }, args.paginas, args.sleep)
        rows = extract_rows(payload)
        prices = [float(r.get("precoUnitario")) for r in rows if r.get("precoUnitario") not in (None, "") and float(r.get("precoUnitario")) > 0]
        rec = {"codigoItemCatalogo": code, "registros": len(rows)}
        if prices:
            rec.update({"min": min(prices), "max": max(prices), "media": round(sum(prices)/len(prices), 4), "mediana": statistics.median(prices)})
            rec["descricaoAmostra"] = rows[0].get("descricaoItem")
        summary_rows.append(rec)
    csv_path = outdir / "resumo_precos_por_codigo.csv"
    write_csv(summary_rows, str(csv_path))
    json_path = outdir / "resumo_precos_por_codigo.json"
    json_path.write_text(json.dumps(summary_rows, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print_json({"xlsx": args.xlsx, "codigos": len(codes), "csv": str(csv_path), "json": str(json_path)})


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="compras-gov", description="CLI para Dados Abertos Compras.gov.br: itens, preços, atas, editais/licitações e PNCP.")
    p.add_argument("--cache", help="diretório de cache local das respostas da API (evita repetir consultas idênticas)")
    sub = p.add_subparsers(dest="cmd", required=True)

    s = sub.add_parser("item", help="consultar catálogo de material por código, classe ou descrição")
    common_page_args(s)
    s.add_argument("--codigo-item", type=int, dest="codigoItem")
    s.add_argument("--descricao", dest="descricaoItem")
    s.add_argument("--codigo-grupo", type=int, dest="codigoGrupo")
    s.add_argument("--codigo-classe", type=int, dest="codigoClasse")
    s.add_argument("--status", type=lambda x: x.lower() in ["1","true","sim","s"], dest="statusItem")
    s.set_defaults(func=lambda a: cmd_endpoint(a, "material-item", vars(a)))

    s = sub.add_parser("precos", help="consultar preços praticados de material por código do catálogo")
    common_page_args(s)
    s.add_argument("--codigo-item", required=True, type=int, dest="codigoItemCatalogo")
    s.add_argument("--inicio", default=date_range_default()[0], dest="dataCompraInicio")
    s.add_argument("--fim", default=date_range_default()[1], dest="dataCompraFim")
    s.add_argument("--uasg", dest="codigoUasg")
    s.add_argument("--uf", dest="estado")
    s.add_argument("--codigo-classe", type=int, dest="codigoClasse")
    s.set_defaults(func=lambda a: cmd_endpoint(a, "precos-material", vars(a)))

    s = sub.add_parser("preco-resumo", help="estatística rápida dos preços praticados por código de item")
    common_page_args(s)
    s.add_argument("--codigo-item", required=True, type=int)
    s.add_argument("--inicio", default=date_range_default()[0])
    s.add_argument("--fim", default=date_range_default()[1])
    s.add_argument("--uasg")
    s.add_argument("--uf")
    s.set_defaults(func=cmd_preco_resumo)

    s = sub.add_parser("atas", help="consultar atas de registro de preços por item, unidade, modalidade ou vigência")
    common_page_args(s)
    s.add_argument("--codigo-item", type=int, dest="codigoItem")
    s.add_argument("--unidade-gerenciadora", type=int, dest="codigoUnidadeGerenciadora")
    s.add_argument("--modalidade", dest="codigoModalidadeCompra")
    s.add_argument("--inicio", required=True, dest="dataVigenciaInicialMin")
    s.add_argument("--fim", required=True, dest="dataVigenciaInicialMax")
    s.add_argument("--numero-compra", dest="numeroCompra")
    s.set_defaults(func=lambda a: cmd_endpoint(a, "atas-item", vars(a)))

    s = sub.add_parser("editais-legado", help="consultar licitações/editais no módulo legado por data de publicação")
    common_page_args(s)
    s.add_argument("--inicio", required=True, dest="data_publicacao_inicial")
    s.add_argument("--fim", required=True, dest="data_publicacao_final")
    s.add_argument("--uasg", type=int)
    s.add_argument("--numero-aviso", type=int, dest="numero_aviso")
    s.add_argument("--modalidade", type=int)
    s.add_argument("--lei14133", type=lambda x: x.lower() in ["1","true","sim","s"], dest="pertence14133")
    s.set_defaults(func=lambda a: cmd_endpoint(a, "licitacoes-legado", vars(a)))

    s = sub.add_parser("contratacoes", help="consultar contratações PNCP Lei 14.133")
    common_page_args(s)
    s.add_argument("--inicio", required=True, dest="dataPublicacaoPncpInicial")
    s.add_argument("--fim", required=True, dest="dataPublicacaoPncpFinal")
    s.add_argument("--modalidade", required=True, type=int, dest="codigoModalidade")
    s.add_argument("--cnpj-orgao", dest="orgaoEntidadeCnpj")
    s.add_argument("--uasg", dest="unidadeOrgaoCodigoUnidade")
    s.add_argument("--uf", dest="unidadeOrgaoUfSigla")
    s.set_defaults(func=lambda a: cmd_endpoint(a, "contratacoes-pncp", vars(a)))

    s = sub.add_parser("itens-contratacao", help="consultar itens de contratações PNCP por código de item/período")
    common_page_args(s)
    s.add_argument("--inicio", required=True, dest="dataInclusaoPncpInicial")
    s.add_argument("--fim", required=True, dest="dataInclusaoPncpFinal")
    s.add_argument("--codigo-item", type=int, dest="codItemCatalogo")
    s.add_argument("--material-ou-servico", dest="materialOuServico")
    s.add_argument("--cnpj-orgao", dest="orgaoEntidadeCnpj")
    s.add_argument("--uasg", dest="unidadeOrgaoCodigoUnidade")
    s.set_defaults(func=lambda a: cmd_endpoint(a, "itens-contratacoes-pncp", vars(a)))

    s = sub.add_parser("resultados", help="consultar resultados homologados PNCP")
    common_page_args(s)
    s.add_argument("--inicio", required=True, dest="dataResultadoPncpInicial")
    s.add_argument("--fim", required=True, dest="dataResultadoPncpFinal")
    s.add_argument("--cnpj-orgao", dest="orgaoEntidadeCnpj")
    s.add_argument("--uasg", dest="unidadeOrgaoCodigoUnidade")
    s.add_argument("--fornecedor", dest="niFornecedor")
    s.set_defaults(func=lambda a: cmd_endpoint(a, "resultados-pncp", vars(a)))

    s = sub.add_parser("planilha-precos", help="extrai códigos de uma planilha DFD e gera resumo de preços por item")
    s.add_argument("xlsx")
    s.add_argument("--inicio", default=date_range_default()[0])
    s.add_argument("--fim", default=date_range_default()[1])
    s.add_argument("--out", default="output/compras-gov-planilha")
    s.add_argument("--tipo", choices=["material", "servico"], default="material", help="módulo de preços a consultar (CATMAT ou CATSER)")
    s.add_argument("--tamanho-pagina", type=int, default=10, dest="tamanhoPagina")
    s.add_argument("--paginas", type=int, default=1)
    s.add_argument("--sleep", type=float, default=0.15)
    s.add_argument("--max-itens", type=int)
    s.set_defaults(func=cmd_planilha_pesquisar)

    s = sub.add_parser("sugerir-codigo", help="sugere códigos CATMAT para uma descrição livre, ranqueados por similaridade")
    s.add_argument("--descricao", required=True, help="descrição do item para busca no catálogo de materiais")
    s.add_argument("--catalogo-pdm", default="output/.cache/catalogo_pdm.json", dest="catalogo_pdm", help="cache local do catálogo PDM (baixado na primeira execução)")
    s.add_argument("--top-pdm", type=int, default=3, dest="top_pdm", help="quantos PDMs candidatos expandir em itens")
    s.add_argument("--limit", type=int, default=10, help="máximo de códigos sugeridos")
    s.add_argument("--sleep", type=float, default=0.1)
    s.add_argument("--csv", help="salvar sugestões em CSV")
    s.set_defaults(func=cmd_sugerir_codigo)

    s = sub.add_parser("endpoints", help="listar endpoints usados pela integração")
    s.set_defaults(func=lambda a: print_json(ENDPOINTS))
    return p


def main(argv: List[str] | None = None) -> int:
    global CACHE_DIR
    parser = build_parser()
    args = parser.parse_args(argv)
    if getattr(args, "cache", None):
        CACHE_DIR = Path(args.cache)
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
    args.func(args)
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
