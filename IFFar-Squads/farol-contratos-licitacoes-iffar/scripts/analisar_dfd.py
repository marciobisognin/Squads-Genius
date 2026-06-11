#!/usr/bin/env python3
import argparse, csv, html, json, math, re, statistics, shutil
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

RISK_COLOR = {'ALTO':'FFFF9999','MÉDIO':'FFFFE699','BAIXO':'FFD9EAD3','OK':'FFE2F0D9'}
PACK_TERMS = ['CAIXA','PACOTE','FARDO','KIT','CONJUNTO','JOGO','PAR','ROLO','FRASCO','POTE','GALÃO','SACO','EMBALAGEM','CARTELA']
RESTRICTIVE = ['MARCA','MODELO EXCLUSIVO','FABRICAÇÃO NACIONAL','NACIONAL']
TYPO = {'CADO':'CABO','INXIDÁVEL':'INOXIDÁVEL','M`NIMA':'MÍNIMA','MINIMA':'MÍNIMA','CARACTERISTICAS':'CARACTERÍSTICAS','ACO ':'AÇO '}


def norm(s):
    return re.sub(r'\s+', ' ', str(s or '').strip()).upper()

def num(v):
    if v is None or v == '': return None
    if isinstance(v, (int,float)): return float(v)
    try: return float(str(v).replace('.','').replace(',','.'))
    except Exception: return None

def find_header(ws):
    for r in range(1, min(ws.max_row,30)+1):
        vals=[norm(ws.cell(r,c).value) for c in range(1, ws.max_column+1)]
        if any('DESCRIÇÃO DO ITEM' in v for v in vals) and any('UNIDADE' in v for v in vals):
            return r
    return 1

def campus_for_col(ws, col, header_row):
    # prefer the nearest non-empty cell on row 2 to the left within a 2-column campus block
    for c in [col, col-1, col-2]:
        if c >= 1:
            v = ws.cell(2,c).value
            if v and 'DOCUMENTO DE FORMALIZAÇÃO' not in str(v).upper() and 'VALOR' not in str(v).upper():
                return str(v).strip()
    return ws.cell(header_row,col).column_letter

def is_yellow(cell):
    rgb = (cell.fill.fgColor.rgb or '').upper()
    return rgb in ('FFFFFF00','FFFF00','00FFFF00')

def detect_columns(ws, header_row):
    cols={}
    for c in range(1, ws.max_column+1):
        h=norm(ws.cell(header_row,c).value)
        if 'CÓDIGO' in h or 'CODIGO' in h: cols['codigo']=c
        elif 'DESCRIÇÃO' in h or 'DESCRICAO' in h: cols['descricao']=c
        elif 'UNIDADE' in h: cols['unidade']=c
        elif 'VALOR ESTIMADO NA ÚLTIMA' in h or ('VALOR ESTIMADO' in h and 'TOTAL' not in h): cols['preco']=c
    qcols=[]
    for c in range(1, ws.max_column+1):
        h=norm(ws.cell(header_row,c).value)
        if 'QUANTIDADE ESTIMADA' in h:
            sample=ws.cell(header_row+1,c)
            if is_yellow(sample) or is_yellow(ws.cell(header_row,c)):
                qcols.append((c, campus_for_col(ws,c,header_row)))
    if not qcols:
        # fallback: quantity columns before total value columns
        for c in range(1, ws.max_column+1):
            h=norm(ws.cell(header_row,c).value)
            if 'QUANTIDADE ESTIMADA' in h:
                qcols.append((c, campus_for_col(ws,c,header_row)))
    return cols,qcols

def description_findings(desc, unidade):
    d=norm(desc); u=norm(unidade)
    findings=[]
    if len(d) < 45:
        findings.append(('ALTO','DESCRIÇÃO','Descrição muito curta; complementar especificação técnica mínima.'))
    if not re.search(r'\d', d):
        findings.append(('MÉDIO','DESCRIÇÃO','Descrição sem medida/capacidade/dimensão numérica; verificar se o item exige especificação objetiva.'))
    for wrong,right in TYPO.items():
        if wrong in d:
            findings.append(('BAIXO','DESCRIÇÃO',f'Possível erro de digitação: "{wrong}"; sugerir "{right}".'))
    for term in RESTRICTIVE:
        if term in d:
            sev='MÉDIO' if term != 'FABRICAÇÃO NACIONAL' else 'ALTO'
            findings.append((sev,'DESCRIÇÃO',f'Termo potencialmente restritivo/direcionador: "{term}"; revisar justificativa ou remover.'))
    if any(t in d for t in ['INOX','ALUMÍNIO','ALUMINIO','VIDRO','PLÁSTICO','PLASTICO','POLIPROPILENO','AÇO','ACO']):
        pass
    else:
        if any(k in d for k in ['PANELA','CAÇAROLA','FORMA','COPO','XÍCARA','XICARA','TALHER','FACA','COLHER','PRATO','GARRAFA','JARRA']):
            findings.append(('MÉDIO','DESCRIÇÃO','Material constitutivo não aparece de forma clara; confirmar especificação.'))
    # unit/package mismatch
    if u in ['UNIDADE','UND','UN'] and any(t in d for t in PACK_TERMS):
        findings.append(('MÉDIO','UNIDADE','Descrição menciona embalagem/conjunto, mas unidade de fornecimento está como unidade; esclarecer se a cotação é por unidade ou embalagem.'))
    if u in ['CAIXA','PACOTE','FARDO'] and not any(t in d for t in PACK_TERMS):
        findings.append(('MÉDIO','UNIDADE',f'Unidade "{u}" exige quantidade por embalagem clara na descrição.'))
    return findings

def outlier_findings(values_by_campus):
    vals=[v for c,v in values_by_campus if v is not None and v>0]
    out=[]
    if len(vals) < 4:
        return out
    med=statistics.median(vals)
    qs=statistics.quantiles(vals, n=4, method='inclusive') if len(vals)>=4 else [med,med,med]
    q1,q3=qs[0],qs[2]
    iqr=max(q3-q1,0)
    mad=statistics.median([abs(v-med) for v in vals]) if vals else 0
    for campus,v in values_by_campus:
        if v is None: continue
        if v > 0:
            high = (iqr>0 and v>q3+1.5*iqr and v>=2.5*max(med,1)) or (mad>0 and abs(v-med)/(1.4826*mad)>3.5 and v>med)
            low = (med>=4 and v<=0.25*med and v<q1-1.5*iqr) or (mad>0 and abs(v-med)/(1.4826*mad)>3.5 and v<med)
            if high:
                out.append(('ALTO','OUTLIER',f'{campus} solicitou {v:g}, acima da mediana {med:g}; confirmar demanda ou possível erro.'))
            elif low:
                out.append(('MÉDIO','OUTLIER',f'{campus} solicitou {v:g}, abaixo da mediana {med:g}; confirmar interpretação do item.'))
    zeros=[campus for campus,v in values_by_campus if v == 0]
    positives=sum(1 for c,v in values_by_campus if v and v>0)
    if positives >= max(6, math.ceil(0.7*len(values_by_campus))) and zeros and len(zeros)<=3:
        out.append(('BAIXO','OUTLIER',f'Campus sem demanda em item com adesão ampla: {", ".join(zeros[:5])}; confirmar se zero é intencional.'))
    return out

def price_findings(price):
    if price is None: return [('MÉDIO','PREÇO','Preço estimado ausente; realizar pesquisa/validação.')]
    if price <= 0: return [('ALTO','PREÇO','Preço estimado zero ou negativo; corrigir antes da licitação.')]
    return []

def risk_rank(risks):
    if 'ALTO' in risks: return 'ALTO'
    if 'MÉDIO' in risks: return 'MÉDIO'
    if 'BAIXO' in risks: return 'BAIXO'
    return 'OK'

def analyze(input_path, outdir):
    input_path=Path(input_path); outdir=Path(outdir); outdir.mkdir(parents=True, exist_ok=True)
    wb=openpyxl.load_workbook(input_path)
    ws=wb[wb.sheetnames[0]]
    header=find_header(ws)
    cols,qcols=detect_columns(ws, header)
    required=['codigo','descricao','unidade']
    missing=[k for k in required if k not in cols]
    if missing: raise SystemExit('Colunas obrigatórias ausentes: '+', '.join(missing))
    action_col=ws.max_column+1
    headers=['Ações Necessárias','Nível de Risco','Tipos de Achado','Outliers Quantitativos','Sugestão de Decisão']
    for i,h in enumerate(headers):
        cell=ws.cell(header, action_col+i, h)
        cell.fill=PatternFill('solid', fgColor='FF1F4E78')
        cell.font=Font(color='FFFFFFFF', bold=True)
        cell.alignment=Alignment(wrap_text=True, vertical='top')
    findings=[]; rows=0; counts={'ALTO':0,'MÉDIO':0,'BAIXO':0,'OK':0}
    type_counts={}
    for r in range(header+1, ws.max_row+1):
        desc=ws.cell(r, cols['descricao']).value
        if not desc or not str(desc).strip(): continue
        rows += 1
        codigo=ws.cell(r, cols.get('codigo',1)).value
        item=ws.cell(r, 1).value
        unidade=ws.cell(r, cols.get('unidade')).value
        price=num(ws.cell(r, cols.get('preco')).value) if cols.get('preco') else None
        vals=[(campus, num(ws.cell(r,c).value)) for c,campus in qcols]
        fs=[]
        fs.extend(description_findings(desc, unidade))
        fs.extend(price_findings(price))
        fs.extend(outlier_findings(vals))
        if not fs:
            actions='Aprovado sem ressalvas'
            risk='OK'; types='OK'; outs=''; decision='Aprovar sem ressalvas automatizadas.'
        else:
            risk=risk_rank([f[0] for f in fs])
            actions=' | '.join(f[2] for f in fs)
            types=', '.join(sorted(set(f[1] for f in fs)))
            outs=' | '.join(f[2] for f in fs if f[1]=='OUTLIER')
            decision='Revisar antes de consolidar.' if risk in ['ALTO','MÉDIO'] else 'Baixa criticidade; revisar se houver tempo.'
        counts[risk]+=1
        for f in fs: type_counts[f[1]]=type_counts.get(f[1],0)+1
        fill=PatternFill('solid', fgColor=RISK_COLOR[risk])
        for i,val in enumerate([actions,risk,types,outs,decision]):
            cell=ws.cell(r, action_col+i, val)
            cell.alignment=Alignment(wrap_text=True, vertical='top')
            if i==1: cell.fill=fill
        if fs:
            for sev,typ,msg in fs:
                findings.append({'linha':r,'item':item,'codigo':codigo,'unidade':unidade,'risco':sev,'tipo':typ,'achado':msg,'descricao':str(desc)[:300]})
    for c in range(action_col, action_col+len(headers)):
        ws.column_dimensions[ws.cell(header,c).column_letter].width = 32 if c==action_col else 22
    audited=outdir/(input_path.stem+'_AUDITADA.xlsx')
    wb.save(audited)
    csv_path=outdir/'achados_auditoria.csv'
    with csv_path.open('w', newline='', encoding='utf-8-sig') as f:
        w=csv.DictWriter(f, fieldnames=['linha','item','codigo','unidade','risco','tipo','achado','descricao'])
        w.writeheader(); w.writerows(findings)
    top=findings[:20]
    report=outdir/'relatorio_executivo.md'
    report.write_text(make_report(input_path, rows, counts, type_counts, findings, qcols), encoding='utf-8')
    dash=outdir/'dashboard_auditoria.html'
    dash.write_text(make_dashboard(input_path, rows, counts, type_counts, findings), encoding='utf-8')
    summary={'input':str(input_path),'items_analisados':rows,'campi_detectados':[c for _,c in qcols],'achados':len(findings),'riscos':counts,'tipos':type_counts,'outputs':{'planilha':str(audited),'csv':str(csv_path),'relatorio':str(report),'dashboard':str(dash)}}
    (outdir/'summary.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    return summary

def make_report(input_path, rows, counts, type_counts, findings, qcols):
    crit=[f for f in findings if f['risco']=='ALTO'][:15]
    lines=[]
    lines.append(f'# Relatório Executivo — Auditoria de DFD/Listas de Itens\n')
    lines.append(f'**Arquivo analisado:** `{Path(input_path).name}`')
    lines.append(f'**Data de geração:** {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    lines.append(f'**Itens analisados:** {rows}')
    lines.append(f'**Campi/colunas oficiais detectados:** {len(qcols)} — '+', '.join(c for _,c in qcols))
    lines.append('\n## Síntese de risco')
    for k in ['ALTO','MÉDIO','BAIXO','OK']:
        lines.append(f'- {k}: {counts.get(k,0)} itens')
    lines.append('\n## Achados por tipo')
    for k,v in sorted(type_counts.items(), key=lambda x:-x[1]): lines.append(f'- {k}: {v}')
    lines.append('\n## Itens prioritários')
    if crit:
        for f in crit:
            lines.append(f'- Linha {f["linha"]}, item {f["item"]}, código {f["codigo"]}: [{f["tipo"]}] {f["achado"]}')
    else:
        lines.append('- Não foram detectados achados de risco alto pelas regras automatizadas.')
    lines.append('\n## Recomendações imediatas')
    lines.append('- Revisar primeiro itens com risco ALTO e achados de OUTLIER/PREÇO/UNIDADE.')
    lines.append('- Validar com o campus demandante qualquer quantitativo muito acima ou abaixo da mediana multicampi.')
    lines.append('- Padronizar descrições saneadas em uma base de conhecimento para reaproveitamento em próximas contratações.')
    lines.append('- Para itens financeiramente relevantes, acionar pesquisa externa de preços em PNCP/Compras.gov/Painel de Preços.')
    lines.append('\n## Possíveis extensões para tomada de decisão')
    lines.append('- Histórico por campus para monitorar recorrência de erros e curva de consumo.')
    lines.append('- Modelo preditivo simples por categoria/campus, usando séries históricas de DFD e execução contratual.')
    lines.append('- Painel de saneamento com status: pendente, confirmado pelo campus, corrigido, aprovado.')
    lines.append('\n## Limitações')
    lines.append('Esta análise é apoio técnico automatizado. A equipe de licitações/contratos deve validar juridicamente, tecnicamente e com os campi os achados antes de alterar o DFD ou edital.')
    lines.append('\nLicença: MIT. Criado por Marcio Bisognin. Instagram: @marciobisognin.')
    return '\n'.join(lines)+'\n'

def make_dashboard(input_path, rows, counts, type_counts, findings):
    cards=''.join(f'<div class="card"><b>{html.escape(k)}</b><span>{v}</span></div>' for k,v in counts.items())
    type_rows=''.join(f'<tr><td>{html.escape(k)}</td><td>{v}</td></tr>' for k,v in sorted(type_counts.items(), key=lambda x:-x[1]))
    finding_rows=''.join(f'<tr><td>{f["linha"]}</td><td>{html.escape(str(f["codigo"]))}</td><td>{html.escape(f["risco"])}</td><td>{html.escape(f["tipo"])}</td><td>{html.escape(f["achado"])}</td></tr>' for f in findings[:80])
    return f"""<!doctype html><html lang=\"pt-BR\"><head><meta charset=\"utf-8\"><title>Dashboard Auditoria DFD</title><style>body{{font-family:Arial,sans-serif;background:#0f172a;color:#e5e7eb;margin:0;padding:28px}}h1{{color:#93c5fd}}.grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}}.card{{background:#1e293b;border:1px solid #334155;border-radius:14px;padding:18px}}.card span{{display:block;font-size:34px;color:#fbbf24}}table{{width:100%;border-collapse:collapse;background:#111827;margin-top:18px}}td,th{{border:1px solid #374151;padding:8px;vertical-align:top}}th{{background:#1f2937}}.muted{{color:#94a3b8}}</style></head><body><h1>Dashboard — Auditoria de DFD/Listas de Itens</h1><p class=\"muted\">Arquivo: {html.escape(Path(input_path).name)} | Itens analisados: {rows}</p><div class=\"grid\">{cards}</div><h2>Achados por tipo</h2><table><tr><th>Tipo</th><th>Quantidade</th></tr>{type_rows}</table><h2>Achados prioritários</h2><table><tr><th>Linha</th><th>Código</th><th>Risco</th><th>Tipo</th><th>Achado</th></tr>{finding_rows}</table><p class=\"muted\">Licença: MIT. Criado por Marcio Bisognin. Instagram: @marciobisognin.</p></body></html>"""

def main():
    ap=argparse.ArgumentParser(description='Audita planilha DFD/lista de itens de licitações e contratos.')
    ap.add_argument('planilha')
    ap.add_argument('--out', default='output/auditoria')
    args=ap.parse_args()
    summary=analyze(args.planilha, args.out)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
if __name__ == '__main__': main()
