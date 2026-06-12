#!/usr/bin/env python3
"""Camada de enriquecimento Compras.gov para o Farol Contratos & Licitações IFFar.

Fluxo:
1. Executa a auditoria DFD existente (`analisar_dfd.py`).
2. Usa o CLI `compras_gov.py planilha-precos` para pesquisar preços praticados.
3. Acrescenta colunas de benchmark Compras.gov à planilha auditada.
4. Gera relatório complementar com cobertura e recomendações.
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from farol_common import similarity

ROOT = Path(__file__).resolve().parents[1]
SIMILARITY_FLOOR = 0.15  # abaixo disso a comparação de preço externo é tratada como pouco confiável
ANALYZER = ROOT / "scripts" / "analisar_dfd.py"
COMPRAS = ROOT / "scripts" / "compras_gov.py"


def default_dates(days: int = 730) -> tuple[str, str]:
    end = date.today()
    start = end - timedelta(days=days)
    return start.isoformat(), end.isoformat()


def run(cmd: List[str]) -> None:
    print("+", " ".join(str(c) for c in cmd), file=sys.stderr)
    subprocess.run(cmd, check=True)


def find_header(ws) -> int:
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for r in range(1, min(max_row, 40) + 1):
        vals = [str(ws.cell(r, c).value or "").upper() for c in range(1, min(max_col, 60) + 1)]
        if any("CÓDIGO" in v or "CODIGO" in v for v in vals) and any("DESCRI" in v for v in vals):
            return r
    return 1


def find_col(ws, header: int, labels: List[str]) -> int | None:
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = str(ws.cell(header, c).value or "").upper()
        if any(label in v for label in labels):
            return c
    return None


def classify_price(estimado: float | None, mediana: float | None) -> str:
    if estimado is None or mediana is None or mediana <= 0:
        return "Sem comparação suficiente"
    ratio = estimado / mediana
    if ratio >= 1.5:
        return f"Acima da mediana Compras.gov ({ratio:.2f}x); revisar estimativa"
    if ratio <= 0.55:
        return f"Abaixo da mediana Compras.gov ({ratio:.2f}x); verificar subestimativa ou especificação diferente"
    return f"Compatível com mediana Compras.gov ({ratio:.2f}x)"


def as_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        try:
            return float(str(v).replace(".", "").replace(",", "."))
        except Exception:
            return None


def enrich_workbook(audited_path: Path, price_summary_path: Path, out_path: Path) -> Dict[str, Any]:
    data = json.loads(price_summary_path.read_text(encoding="utf-8"))
    by_code = {int(row["codigoItemCatalogo"]): row for row in data if row.get("codigoItemCatalogo") is not None}
    wb = openpyxl.load_workbook(audited_path)
    ws = wb[wb.sheetnames[0]]
    header = find_header(ws)
    code_col = find_col(ws, header, ["CÓDIGO", "CODIGO"])
    price_col = find_col(ws, header, ["VALOR ESTIMADO NA ÚLTIMA", "VALOR ESTIMADO"])
    desc_col = find_col(ws, header, ["DESCRI"])
    if not code_col:
        raise SystemExit("Não encontrei coluna CÓDIGO na planilha auditada.")
    start = (ws.max_column or 1) + 1
    headers = [
        "Compras.gov Registros",
        "Compras.gov Preço Mínimo",
        "Compras.gov Preço Médio",
        "Compras.gov Preço Mediano",
        "Compras.gov Preço Máximo",
        "Compras.gov Descrição Amostra",
        "Compras.gov Similaridade Descrição",
        "Avaliação Preço x Compras.gov",
    ]
    fill = PatternFill("solid", fgColor="FF38761D")
    for i, h in enumerate(headers):
        cell = ws.cell(header, start + i, h)
        cell.fill = fill
        cell.font = Font(color="FFFFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[cell.column_letter].width = 28 if i != 5 else 48
    enriched = 0
    comparable = 0
    alerts = 0
    low_similarity = 0
    for r in range(header + 1, (ws.max_row or header) + 1):
        try:
            code = int(float(ws.cell(r, code_col).value))
        except Exception:
            continue
        rec = by_code.get(code)
        if not rec:
            continue
        enriched += 1
        mediana = as_float(rec.get("mediana"))
        estimado = as_float(ws.cell(r, price_col).value) if price_col else None
        descricao_interna = ws.cell(r, desc_col).value if desc_col else None
        sim = similarity(descricao_interna, rec.get("descricaoAmostra")) if rec.get("descricaoAmostra") else None
        avaliacao = classify_price(estimado, mediana)
        if sim is not None and sim < SIMILARITY_FLOOR and "mediana" in avaliacao:
            avaliacao += " — ATENÇÃO: baixa equivalência de descrição; comparar com cautela"
            low_similarity += 1
        elif "Acima" in avaliacao or "Abaixo" in avaliacao:
            alerts += 1
        if estimado is not None and mediana is not None:
            comparable += 1
        vals = [
            rec.get("registros"), rec.get("min"), rec.get("media"), rec.get("mediana"), rec.get("max"), rec.get("descricaoAmostra"),
            (f"{sim:.0%}" if sim is not None else ""), avaliacao
        ]
        for i, val in enumerate(vals):
            cell = ws.cell(r, start + i, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(out_path)
    return {"itens_enriquecidos": enriched, "comparaveis": comparable, "alertas_preco": alerts, "baixa_similaridade": low_similarity, "saida": str(out_path)}


def write_report(outdir: Path, audit_summary: Dict[str, Any], compras_summary: List[Dict[str, Any]], enrichment: Dict[str, Any], inicio: str, fim: str) -> Path:
    path = outdir / "relatorio_compras_gov.md"
    with path.open("w", encoding="utf-8") as f:
        f.write("# Relatório complementar — Integração Compras.gov\n\n")
        f.write(f"Período pesquisado: **{inicio} a {fim}**\n\n")
        f.write("## Cobertura\n\n")
        f.write(f"- Itens auditados no DFD: {audit_summary.get('items_analisados')}\n")
        f.write(f"- Códigos pesquisados no Compras.gov: {len(compras_summary)}\n")
        f.write(f"- Itens com dados externos encontrados: {enrichment.get('itens_enriquecidos')}\n")
        f.write(f"- Itens comparáveis com preço estimado/mediana: {enrichment.get('comparaveis')}\n")
        f.write(f"- Alertas de preço por comparação externa: {enrichment.get('alertas_preco')}\n")
        f.write(f"- Comparações marcadas com baixa equivalência de descrição: {enrichment.get('baixa_similaridade', 0)}\n\n")
        f.write("## Recomendações de uso\n\n")
        f.write("- Usar a mediana Compras.gov como referência robusta inicial, não como preço final automático.\n")
        f.write("- Quando a descrição do DFD divergir da descrição amostra do Compras.gov, revisar especificação antes de comparar valores.\n")
        f.write("- Itens acima de 1,5x ou abaixo de 0,55x da mediana externa devem receber pesquisa complementar e validação técnica.\n")
        f.write("- Para itens estratégicos, complementar com atas vigentes e contratações PNCP similares.\n\n")
        f.write("Licença: MIT. Criado por Marcio Bisognin. Instagram: @marciobisognin.\n")
    return path


def main() -> int:
    start, end = default_dates()
    ap = argparse.ArgumentParser(description="Audita DFD e enriquece com preços praticados do Compras.gov.")
    ap.add_argument("planilha")
    ap.add_argument("--out", default="output/farol-compras-gov")
    ap.add_argument("--inicio", default=start)
    ap.add_argument("--fim", default=end)
    ap.add_argument("--paginas", type=int, default=1)
    ap.add_argument("--tamanho-pagina", type=int, default=10)
    ap.add_argument("--sleep", type=float, default=0.15)
    ap.add_argument("--max-itens", type=int, help="limita códigos pesquisados para teste/demonstração")
    args = ap.parse_args()

    outdir = Path(args.out)
    audit_dir = outdir / "01_auditoria_dfd"
    compras_dir = outdir / "02_compras_gov"
    outdir.mkdir(parents=True, exist_ok=True)

    run([sys.executable, str(ANALYZER), args.planilha, "--out", str(audit_dir)])
    cache_dir = outdir / ".cache"
    cmd = [sys.executable, str(COMPRAS), "--cache", str(cache_dir), "planilha-precos", args.planilha, "--inicio", args.inicio, "--fim", args.fim, "--out", str(compras_dir), "--paginas", str(args.paginas), "--tamanho-pagina", str(args.tamanho_pagina), "--sleep", str(args.sleep)]
    if args.max_itens:
        cmd += ["--max-itens", str(args.max_itens)]
    run(cmd)

    audit_summary = json.loads((audit_dir / "summary.json").read_text(encoding="utf-8"))
    compras_summary_path = compras_dir / "resumo_precos_por_codigo.json"
    compras_summary = json.loads(compras_summary_path.read_text(encoding="utf-8"))
    audited_path = Path(audit_summary["outputs"]["planilha"])
    enriched_path = outdir / (Path(args.planilha).stem + "_AUDITADA_COMPRAS_GOV.xlsx")
    enrichment = enrich_workbook(audited_path, compras_summary_path, enriched_path)
    report_path = write_report(outdir, audit_summary, compras_summary, enrichment, args.inicio, args.fim)
    summary = {
        "planilha_original": args.planilha,
        "periodo_compras_gov": {"inicio": args.inicio, "fim": args.fim},
        "auditoria_dfd": audit_summary,
        "compras_gov": {"codigos_pesquisados": len(compras_summary), "arquivo_json": str(compras_summary_path), "arquivo_csv": str(compras_dir / "resumo_precos_por_codigo.csv")},
        "enriquecimento": enrichment,
        "relatorio_compras_gov": str(report_path),
    }
    (outdir / "summary_compras_gov.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
